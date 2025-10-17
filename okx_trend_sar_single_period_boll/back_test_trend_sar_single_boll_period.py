#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 添加当前目录到Python路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from trend_sar_single_period_boll_strategy import TrendSarStrategy
from database_config import LOCAL_DATABASE_CONFIG
from database_service import DatabaseService
from strategy_configs import get_strategy_config, print_config_info
from performance_analysis import PerformanceAnalyzer

def create_output_directory(config, annual_return):
    """
    创建统一的输出目录结构
    /back_test_data/{币种}/{年化收益率}-{时间戳}/
    
    Args:
        config: 策略配置
        annual_return: 年化收益率（用于文件夹命名）
    
    Returns:
        str: 输出目录路径
    """
    # 获取币种
    long_coin = config['long_coin']
    
    # 生成时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 格式化年化收益率（保留2位小数）
    annual_return_str = f"{annual_return:+.2f}%"
    
    # 创建目录结构
    base_dir = "back_test_data"
    coin_dir = os.path.join(base_dir, long_coin)
    result_dir = os.path.join(coin_dir, f"{annual_return_str}-{timestamp}")
    
    # 确保目录存在
    os.makedirs(result_dir, exist_ok=True)
    
    print(f"📁 创建输出目录: {result_dir}")
    return result_dir

def export_trades_to_excel(trades, config, output_dir):
    """
    导出交易记录到Excel文件
    
    Args:
        trades: 交易记录列表
        config: 策略配置
        output_dir: 输出目录
    """
    if not trades:
        print("❌ 没有交易记录，跳过Excel导出")
        return None
    
    # 生成文件名（去掉时间戳）
    filename = os.path.join(output_dir, "交易记录.xlsx")
    
    # 合并开仓和平仓记录，按时间顺序
    trade_data = []
    current_entry = None  # 记录当前开仓信息
    
    for i, trade in enumerate(trades):
        if trade['signal_type'] in ['OPEN_LONG', 'OPEN_SHORT']:
            # 开仓记录
            direction = "做多" if trade['signal_type'] == 'OPEN_LONG' else "做空"
            stop_loss = trade.get('stop_loss', 'N/A')
            take_profit = trade.get('take_profit', 'N/A')
            
            # 🔄 获取实际投入金额（来自策略传递的复投信息）
            trade_amount = trade.get('invested_amount', 0)
            
            # 调试信息：检查开仓信号是否包含投入金额
            if trade_amount == 0:
                print(f"⚠️  警告：开仓信号 {trade['signal_type']} 缺少 invested_amount 信息")
                print(f"    信号内容: {trade}")
                # 🔄 使用当前现金余额计算（复投逻辑）
                position_size = config.get('position_size_percentage', 100) / 100
                # 单账户模式：使用统一现金余额
                current_balance = trade.get('cash_balance', config['initial_capital'])
                trade_amount = current_balance * position_size
                print(f"    🔄 复投计算: 当前余额${current_balance:,.2f} × {position_size*100}% = ${trade_amount:,.2f}")
            
            # 单账户模式：获取统一现金余额信息
            cash_balance = trade.get('cash_balance', 0)
            
            # 保存当前开仓信息，用于平仓时引用
            current_entry = {
                'direction': direction,
                'price': trade['price'],
                'stop_loss': stop_loss,
                'take_profit': take_profit,
                'amount': trade_amount,
                'cash_balance': cash_balance
            }
            
            trade_data.append({
                '序号': len(trade_data) + 1,
                '操作类型': '开仓',
                '交易方向': direction,
                '时间': trade['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if trade.get('timestamp') else 'N/A',
                '价格': f"${trade['price']:.2f}",
                '交易金额': f"${trade_amount:,.2f}",
                '交易份额': f"{trade.get('position_shares', 0):.4f}",
                '止损位': f"${stop_loss:.2f}" if isinstance(stop_loss, (int, float)) else stop_loss,
                '止盈位': f"${take_profit:.2f}" if isinstance(take_profit, (int, float)) else take_profit,
                '盈亏金额': '-',
                '手续费': f"${trade.get('transaction_fee', 0):.2f}",
                '收益率': '-',
                '交易结果': '-',
                '现金余额': f"${cash_balance:,.2f}",
                '原因': trade['reason']
            })
            
        elif trade['signal_type'] in ['STOP_LOSS_LONG', 'STOP_LOSS_SHORT', 'TAKE_PROFIT_LONG', 'TAKE_PROFIT_SHORT', 
                                       'MA_PROFIT_LONG', 'MA_LOSS_LONG', 'MA_PROFIT_SHORT', 'MA_LOSS_SHORT',
                                       'MAX_STOP_LOSS_LONG', 'MAX_STOP_LOSS_SHORT']:
            # 平仓记录
            direction = "做多" if 'LONG' in trade['signal_type'] else "做空"
            profit_loss = trade.get('profit_loss', 0)
            
            # 判断平仓类型和结果 - 更精确地描述平仓原因
            if 'TAKE_PROFIT' in trade['signal_type']:
                close_type = "固定止盈"
                result = "盈利"
                print(f"    🎯 平仓分类: {trade['signal_type']} -> {close_type} ({result})")
            elif 'MA_PROFIT' in trade['signal_type'] or 'MA_LOSS' in trade['signal_type']:
                # 回归MA卖出信号
                if profit_loss > 0:
                    close_type = "回归MA盈利"
                    result = "盈利"
                    print(f"    📊 平仓分类: {trade['signal_type']} + 盈利${profit_loss:.2f} -> {close_type} ({result})")
                else:
                    close_type = "回归MA亏损"
                    result = "亏损"
                    print(f"    📊 平仓分类: {trade['signal_type']} + 亏损${profit_loss:.2f} -> {close_type} ({result})")
            elif 'MAX_STOP_LOSS' in trade['signal_type']:
                # 最大固定止损
                close_type = "最大固定止损"
                result = "亏损"
                print(f"    🔒 平仓分类: {trade['signal_type']} + 亏损${profit_loss:.2f} -> {close_type} ({result})")
            else:  # STOP_LOSS - 基于固定比例或布林带宽度的止损
                if profit_loss > 0:
                    close_type = "动态止盈"  # 其他原因导致的盈利平仓
                    result = "盈利"
                    print(f"    📈 平仓分类: {trade['signal_type']} + 盈利${profit_loss:.2f} -> {close_type} ({result})")
                else:
                    close_type = "动态止损"  # 固定比例或布林带止损
                    result = "亏损"
                    print(f"    📉 平仓分类: {trade['signal_type']} + 亏损${profit_loss:.2f} -> {close_type} ({result})")
            
            # 计算收益率（基于投入本金）
            if profit_loss != 0 and current_entry and current_entry.get('amount', 0) > 0:
                # 使用投入的资金作为基准计算收益率
                invested_capital = current_entry['amount']
                return_rate = (profit_loss / invested_capital) * 100
            else:
                return_rate = 0
            
            # 使用开仓时的止损止盈位和最新现金余额
            if current_entry:
                entry_stop_loss = current_entry['stop_loss']
                entry_take_profit = current_entry['take_profit']
                entry_amount = current_entry['amount']
            else:
                # 如果没有开仓记录，使用默认值
                print(f"⚠️  警告：平仓记录 {trade['signal_type']} 没有对应的开仓记录")
                entry_stop_loss = 'N/A'
                entry_take_profit = 'N/A'
                entry_amount = trade.get('invested_amount', 0)
                
            # 确保 entry_amount 不为0，避免除零错误
            if entry_amount == 0:
                print(f"⚠️  警告：交易金额为0，设置为默认值1000")
                entry_amount = 1000  # 设置一个默认值避免除零错误
            
            # 🔄 获取平仓后的现金余额（单账户模式）
            new_cash_balance = trade.get('new_balance', 0)
            
            # 🔄 平仓交易金额 = 本金 + 盈亏
            close_trade_amount = entry_amount + profit_loss
            print(f"    💰 平仓金额计算: 本金${entry_amount:,.2f} + 盈亏${profit_loss:+.2f} = ${close_trade_amount:,.2f}")
            
            trade_data.append({
                '序号': len(trade_data) + 1,
                '操作类型': f'平仓({close_type})',
                '交易方向': direction,
                '时间': trade['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if trade.get('timestamp') else 'N/A',
                '价格': f"${trade['price']:.2f}",
                '交易金额': f"${close_trade_amount:,.2f}",
                '交易份额': f"{trade.get('position_shares', 0):.4f}",
                '止损位': f"${entry_stop_loss:.2f}" if isinstance(entry_stop_loss, (int, float)) else entry_stop_loss,
                '止盈位': f"${entry_take_profit:.2f}" if isinstance(entry_take_profit, (int, float)) else entry_take_profit,
                '盈亏金额': f"${profit_loss:.2f}",
                '手续费': f"${trade.get('transaction_fee', 0):.2f}",
                '收益率': f"{return_rate:+.2f}%",
                '交易结果': result,
                '现金余额': f"${new_cash_balance:,.2f}",
                '原因': trade['reason']
            })
            
            # 清除当前开仓信息
            current_entry = None
    
    # 创建DataFrame
    df = pd.DataFrame(trade_data)
    
    # 写入Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 交易明细表
        df.to_excel(writer, sheet_name='交易明细', index=False)
        
        # 获取工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['交易明细']
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头样式
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # 设置数据行样式
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = border
                
                # 根据操作类型和交易结果设置颜色
                if cell.column == 2:  # 操作类型列
                    if cell.value == "开仓":
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        cell.font = Font(color="0066CC")
                    elif "平仓" in str(cell.value):
                        if "止盈" in str(cell.value):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif "止损" in str(cell.value):
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                elif cell.column == 6:  # 交易金额列
                    cell.font = Font(bold=True)
                    if "开仓" in worksheet.cell(row=cell.row, column=2).value:
                        cell.font = Font(color="0066CC", bold=True)
                elif cell.column == 11:  # 交易结果列（位置变了，因为加了现金余额列）
                    if cell.value == "盈利":
                        cell.font = Font(color="006100", bold=True)
                    elif cell.value == "亏损":
                        cell.font = Font(color="9C0006", bold=True)
                elif cell.column == 12:  # 现金余额列
                    cell.font = Font(bold=True, color="FF6600")  # 橙色加粗显示现金余额
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 添加统计信息表
        stats_data = []
        if trade_data:
            # 分离开仓和平仓操作进行统计
            open_trades = [t for t in trade_data if t['操作类型'] == '开仓']
            close_trades = [t for t in trade_data if '平仓' in t['操作类型']]
            
            total_complete_trades = len(close_trades)  # 完整交易次数
            profit_trades = len([t for t in close_trades if t['交易结果'] == '盈利'])
            loss_trades = len([t for t in close_trades if t['交易结果'] == '亏损'])
            win_rate = (profit_trades / total_complete_trades) * 100 if total_complete_trades > 0 else 0
            
            # 计算总盈亏（只计算平仓记录）
            total_pnl = sum([float(t['盈亏金额'].replace('$', '')) for t in close_trades if t['盈亏金额'] != '-'])
            
            # 统计开仓类型
            long_opens = len([t for t in open_trades if t['交易方向'] == '做多'])
            short_opens = len([t for t in open_trades if t['交易方向'] == '做空'])
            
            # 统计平仓类型
            take_profit_closes = len([t for t in close_trades if '止盈' in t['操作类型']])
            stop_loss_closes = len([t for t in close_trades if '止损' in t['操作类型']])
            ma_closes = len([t for t in close_trades if '回归MA' in t['操作类型']])
            
            stats_data = [
                ['交易统计', ''],
                ['总开仓次数', len(open_trades)],
                ['总平仓次数', len(close_trades)],
                ['完整交易次数', total_complete_trades],
                ['做多开仓', long_opens],
                ['做空开仓', short_opens],
                ['止盈平仓', take_profit_closes],
                ['止损平仓', stop_loss_closes],
                ['回归MA平仓', ma_closes],
                ['盈利次数', profit_trades],
                ['亏损次数', loss_trades],
                ['胜率', f"{win_rate:.2f}%"],
                ['总盈亏', f"${total_pnl:.2f}"],
                ['', ''],
                ['策略参数', ''],
                ['单周期模式', config['timeframe']],
                ['EMA周期', config['length']],
                ['SAR起始值', config['sar_start']],
                ['SAR递增值', config['sar_increment']],
                ['SAR最大值', config['sar_maximum']],
                ['回归阻尼', config['damping']],
                ['标准差倍数', config['mult']],
                ['固定止盈', f"{config['fixed_take_profit_pct']}%"],
                ['止损策略', 'SAR动态止损跟随'],
                ['开仓机制', 'SAR方向改变时开仓']
            ]
        
        stats_df = pd.DataFrame(stats_data, columns=['项目', '数值'])
        stats_df.to_excel(writer, sheet_name='统计汇总', index=False)
        
        # 设置统计表样式
        stats_sheet = writer.sheets['统计汇总']
        for cell in stats_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        for column in stats_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            stats_sheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ 交易记录已导出到: {filename}")
    open_count = len([t for t in trade_data if t['操作类型'] == '开仓'])
    close_count = len([t for t in trade_data if '平仓' in t['操作类型']])
    print(f"📊 共导出 {len(trade_data)} 条记录 (开仓: {open_count}, 平仓: {close_count})")
    
    return filename

def create_chart_html(chart_data, output_dir):
    """创建交互式图表HTML"""
    import json
    config_info = chart_data.get('config_info', {})
    timeframe = config_info.get('timeframe', '30m')
    
    html_content = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>双周期SAR策略回测 - 真实数据交互图表</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <style>
        body {{
            margin: 0;
            padding: 20px;
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
            color: #ffffff;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1800px;
            margin: 0 auto;
        }}
        .header {{
            text-align: center;
            margin-bottom: 30px;
            background: rgba(255, 255, 255, 0.08);
            padding: 25px;
            border-radius: 15px;
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .header h1 {{
            margin: 0 0 15px 0;
            color: #00d4ff;
            text-shadow: 0 0 30px rgba(0, 212, 255, 0.4);
            font-size: 2.2em;
        }}
        .config-info {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-top: 20px;
            padding: 20px;
            background: rgba(0, 0, 0, 0.2);
            border-radius: 10px;
        }}
        .config-item {{
            display: flex;
            justify-content: space-between;
            padding: 8px 12px;
            background: rgba(255, 255, 255, 0.05);
            border-radius: 6px;
            border-left: 3px solid #00d4ff;
        }}
        .chart-container {{
            width: 100%;
            height: 800px;
            background: rgba(255, 255, 255, 0.08);
            border-radius: 15px;
            padding: 20px;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.4);
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .legend {{
            display: flex;
            justify-content: center;
            gap: 30px;
            margin-bottom: 25px;
            flex-wrap: wrap;
        }}
        .legend-item {{
            display: flex;
            align-items: center;
            gap: 10px;
            font-size: 15px;
            background: rgba(255, 255, 255, 0.08);
            padding: 10px 15px;
            border-radius: 8px;
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .legend-color {{
            width: 16px;
            height: 16px;
            border-radius: 4px;
            box-shadow: 0 0 10px rgba(255, 255, 255, 0.3);
        }}
        .info-panel {{
            background: rgba(255, 255, 255, 0.08);
            padding: 25px;
            border-radius: 15px;
            margin-top: 30px;
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 15px;
        }}
        .info-row {{
            display: flex;
            justify-content: space-between;
            padding: 12px 15px;
            background: rgba(0, 0, 0, 0.3);
            border-radius: 8px;
            border-left: 4px solid #00d4ff;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎯 单周期SAR策略回测 - {timeframe}K线图表</h1>
            <p style="color: #cccccc; font-size: 16px;">💫 {timeframe}SAR动态止损 + 固定止盈 | 🎯 真实回测数据 | 📊 交互式分析</p>
            
            <div class="config-info">
                <div class="config-item">
                    <span>📅 回测开始:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('start_date', 'N/A')}</span>
                </div>
                <div class="config-item">
                    <span>📅 回测结束:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('end_date', 'N/A')}</span>
                </div>
                <div class="config-item">
                    <span>🎯 时间周期:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('timeframe', 'N/A')}</span>
                </div>
                <div class="config-item">
                    <span>⚡ SAR参数:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('sar_start', '0.02')}/{config_info.get('sar_increment', '0.02')}/{config_info.get('sar_maximum', '0.2')}</span>
                </div>
                <div class="config-item">
                    <span>💰 初始资金:</span>
                    <span style="color: #00d4ff; font-weight: bold;">${config_info.get('initial_capital', 0):,}</span>
                </div>
                <div class="config-item">
                    <span>📊 K线数量:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{len(chart_data.get('klineData', []))}</span>
                </div>
                <div class="config-item">
                    <span>🎯 交易信号:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{len(chart_data.get('tradeSignals', []))}</span>
                </div>
            </div>
        </div>

        <div class="legend">
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #00da3c, #00ff41);"></div>
                <span>📈 上涨K线</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #ec0000, #ff1744);"></div>
                <span>📉 下跌K线</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #ffaa00, #ffc107);"></div>
                <span>🎯 {timeframe} SAR (动态止损)</span>
            </div>
        </div>

        <div class="chart-container">
            <div id="main-chart" style="width: 100%; height: 100%;"></div>
        </div>

        <div class="info-panel">
            <h3>📊 当前K线详情</h3>
            <div class="info-grid" id="currentInfo">
                <div class="info-row">
                    <span>🕐 时间:</span>
                    <span id="currentTime">点击K线查看详情</span>
                </div>
                <div class="info-row">
                    <span>📈 开盘:</span>
                    <span id="currentOpen">-</span>
                </div>
                <div class="info-row">
                    <span>⬆️ 最高:</span>
                    <span id="currentHigh">-</span>
                </div>
                <div class="info-row">
                    <span>⬇️ 最低:</span>
                    <span id="currentLow">-</span>
                </div>
                <div class="info-row">
                    <span>📉 收盘:</span>
                    <span id="currentClose">-</span>
                </div>
                <div class="info-row">
                    <span>🎯 {timeframe} SAR:</span>
                    <span id="currentSAR">-</span>
                </div>
            </div>
        </div>
    </div>

    <script>
        const backtestData = {json.dumps(chart_data, ensure_ascii=False, indent=8)};
        
        const chartDom = document.getElementById('main-chart');
        const myChart = echarts.init(chartDom, 'dark');
        
        function getChartOption(data) {{
            return {{
                animation: true,
                backgroundColor: 'transparent',
                tooltip: {{
                    trigger: 'axis',
                    axisPointer: {{
                        type: 'cross',
                        lineStyle: {{
                            color: '#00d4ff',
                            width: 1.5,
                            opacity: 0.8
                        }}
                    }},
                    backgroundColor: 'rgba(20, 20, 30, 0.95)',
                    borderColor: '#00d4ff',
                    borderWidth: 2,
                    textStyle: {{
                        color: '#fff',
                        fontSize: 14
                    }},
                    formatter: function (params) {{
                        const dataIndex = params[0].dataIndex;
                        const time = data.timeData[dataIndex];
                        const kline = data.klineData[dataIndex];
                        const sar = data.sarData30m[dataIndex];
                        
                        updateInfoPanel(time, kline, sar);
                        
                        const change = kline[1] - kline[0];
                        const changePercent = (change / kline[0] * 100).toFixed(3);
                        const changeColor = change >= 0 ? '#00da3c' : '#ec0000';
                        
                        return `
                            <div style="padding: 15px; min-width: 280px;">
                                <div style="margin-bottom: 12px; font-weight: bold; color: #00d4ff; font-size: 16px;">🕐 ${{time}}</div>
                                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 12px;">
                                    <div>📈 开盘: <span style="color: #fff; font-weight: bold;">$${{kline[0].toFixed(2)}}</span></div>
                                    <div>📉 收盘: <span style="color: ${{changeColor}}; font-weight: bold;">$${{kline[1].toFixed(2)}}</span></div>
                                    <div>⬆️ 最高: <span style="color: #00da3c; font-weight: bold;">$${{kline[3].toFixed(2)}}</span></div>
                                    <div>⬇️ 最低: <span style="color: #ec0000; font-weight: bold;">$${{kline[2].toFixed(2)}}</span></div>
                                </div>
                                <div style="margin-bottom: 12px; padding: 8px; background: rgba(0, 212, 255, 0.15); border-radius: 6px; text-align: center;">
                                    📊 涨跌: <span style="color: ${{changeColor}}; font-weight: bold; font-size: 15px;">$${{change.toFixed(2)}} (${{changePercent}}%)</span>
                                </div>
                                <hr style="margin: 12px 0; border-color: #444;">
                                <div style="display: grid; grid-template-columns: 1fr; gap: 8px;">
                                    <div style="color: #ffaa00; font-size: 14px;">🎯 {timeframe} SAR: <span style="font-weight: bold;">$${{sar}}</span></div>
                                </div>
                            </div>
                        `;
                    }}
                }},
                grid: {{
                    left: '6%',
                    right: '6%',
                    bottom: 120,
                    top: 60,
                    backgroundColor: 'transparent'
                }},
                xAxis: {{
                    type: 'category',
                    data: data.timeData,
                    scale: true,
                    boundaryGap: false,
                    axisLine: {{ 
                        onZero: false,
                        lineStyle: {{ color: '#555' }}
                    }},
                    splitLine: {{ 
                        show: true,
                        lineStyle: {{ color: '#333', opacity: 0.6 }}
                    }},
                    axisLabel: {{
                        formatter: function (value) {{
                            return value.slice(5, 16);
                        }},
                        color: '#cccccc',
                        fontSize: 12
                    }}
                }},
                yAxis: {{
                    scale: true,
                    splitArea: {{
                        show: true,
                        areaStyle: {{
                            color: [['rgba(0, 212, 255, 0.03)', 'rgba(0, 212, 255, 0.01)']]
                        }}
                    }},
                    splitLine: {{
                        lineStyle: {{ color: '#333', opacity: 0.6 }}
                    }},
                    axisLabel: {{
                        formatter: function (value) {{
                            return '$' + value.toFixed(0);
                        }},
                        color: '#cccccc',
                        fontSize: 12
                    }},
                    axisLine: {{
                        lineStyle: {{ color: '#555' }}
                    }}
                }},
                dataZoom: [
                    {{
                        type: 'inside',
                        start: 80,
                        end: 100
                    }},
                    {{
                        show: true,
                        type: 'slider',
                        top: '94%',
                        start: 80,
                        end: 100,
                        backgroundColor: 'rgba(45, 45, 45, 0.8)',
                        borderColor: '#00d4ff',
                        fillerColor: 'rgba(0, 212, 255, 0.25)',
                        handleStyle: {{
                            color: '#00d4ff',
                            borderColor: '#00d4ff'
                        }},
                        textStyle: {{
                            color: '#cccccc'
                        }}
                    }}
                ],
                series: [
                    {{
                        name: 'K线',
                        type: 'candlestick',
                        data: data.klineData,
                        itemStyle: {{
                            color: '#00da3c',
                            color0: '#ec0000',
                            borderColor: '#00da3c',
                            borderColor0: '#ec0000',
                            borderWidth: 1.5
                        }}
                    }},
                    {{
                        name: '{timeframe} SAR',
                        type: 'line',
                        data: data.sarData30m,
                        smooth: false,
                        lineStyle: {{
                            color: '#ffaa00',
                            width: 3
                        }},
                        symbol: 'circle',
                        symbolSize: 6,
                        itemStyle: {{
                            color: '#ffaa00',
                            borderColor: '#fff',
                            borderWidth: 1.5
                        }}
                    }}
                ]
            }};
        }}
        
        function updateInfoPanel(time, kline, sar) {{
            document.getElementById('currentTime').textContent = time;
            document.getElementById('currentOpen').textContent = `$${{kline[0].toFixed(2)}}`;
            document.getElementById('currentHigh').textContent = `$${{kline[3].toFixed(2)}}`;
            document.getElementById('currentLow').textContent = `$${{kline[2].toFixed(2)}}`;
            document.getElementById('currentClose').textContent = `$${{kline[1].toFixed(2)}}`;
            document.getElementById('currentSAR').textContent = `$${{sar}}`;
        }}
        
        myChart.setOption(getChartOption(backtestData));
        
        myChart.on('click', function (params) {{
            if (params.componentType === 'series') {{
                const dataIndex = params.dataIndex;
                const time = backtestData.timeData[dataIndex];
                const kline = backtestData.klineData[dataIndex];
                const sar = backtestData.sarData30m[dataIndex];
                
                updateInfoPanel(time, kline, sar);
            }}
        }});
        
        window.addEventListener('resize', function() {{
            myChart.resize();
        }});
        
        console.log('🚀 真实回测数据图表已加载');
    </script>
</body>
</html>'''
    
    # 保存HTML文件到指定目录
    html_file = os.path.join(output_dir, "交互式图表.html")
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"🌐 交互式图表已保存: {html_file}")
    return html_file

def main():
    print("开始测试SAR策略...")
    
    # 使用配置文件中的数据库参数初始化db_service
    db_service = DatabaseService(**LOCAL_DATABASE_CONFIG)

    # 获取策略配置
    config = get_strategy_config()
    print(f"🔍 获取到的配置类型: {type(config)}")
    print(f"🔍 配置中的钉钉webhook: {config.get('dingtalk_webhook')}")
    print(f"🔍 配置中的钉钉secret: {config.get('dingtalk_secret')}")

    # 显示配置信息
    print_config_info()
    
    # 调试钉钉配置
    dingtalk_webhook = config.get('dingtalk_webhook')
    dingtalk_secret = config.get('dingtalk_secret')
    print(f"🔍 回测脚本调试: dingtalk_webhook={dingtalk_webhook}")
    print(f"🔍 回测脚本调试: dingtalk_secret={dingtalk_secret}")
    
    # 创建单周期SAR策略实例
    strategy = TrendSarStrategy(
        timeframe=config['timeframe'],
        length=config['length'],
        damping=config['damping'],
        sar_start=config['sar_start'],
        sar_increment=config['sar_increment'],
        sar_maximum=config['sar_maximum'],
        mult=config['mult'],
        initial_capital=config['initial_capital'],
        position_size_percentage=config['position_size_percentage'],
        fixed_take_profit_pct=config['fixed_take_profit_pct'],
        max_loss_pct=config['max_loss_pct'],
        volatility_timeframe=config['volatility_timeframe'],
        volatility_length=config['volatility_length'],
        volatility_mult=config['volatility_mult'],
        volatility_ema_period=config['volatility_ema_period'],
        volatility_threshold=config['volatility_threshold'],
        basis_change_threshold=config['basis_change_threshold'],
        dingtalk_webhook=dingtalk_webhook,
        dingtalk_secret=dingtalk_secret
    )

    print(f"\n单周期SAR策略初始化完成")
    print(f"📊 时间周期: {config['timeframe']}")
    print(f"🔧 SAR参数: start={config['sar_start']}, increment={config['sar_increment']}, maximum={config['sar_maximum']}")
    print(f"📈 EMA周期: {config['length']}")
    print(f"💰 回归阻尼: {config['damping']} | 标准差倍数: {config['mult']}")
    print(f"🎯 交易逻辑: SAR方向改变开仓 → 固定止盈{config['fixed_take_profit_pct']}% + SAR动态止损")
    
    # 交易统计
    trades = []
    initial_capital = config['initial_capital']
    current_position = None

    df = db_service.get_kline_data(config['long_coin'], config['start_date'], config['end_date'])

    if df.empty:
        print("未获取到数据，请检查数据库连接和表结构")
        return

    print(f"\n📊 获取到 {len(df)} 个数据点")
    print(f"时间范围: {df['timestamp'].min()} 到 {df['timestamp'].max()}")
    print(f"价格范围: ${df['low'].min():.2f} - ${df['high'].max():.2f}")

    # 🔥 滤波器预热：获取回测开始前的历史数据
    start_timestamp = pd.to_datetime(config['start_date'])
    warmup_days = 60  # 🔥 增加到180天的数据预热（确保EMA和SAR指标完全稳定，接近TradingView效果）
    warmup_start = start_timestamp - pd.Timedelta(days=warmup_days)
    warmup_start_str = warmup_start.strftime('%Y-%m-%d %H:%M:%S')
    warmup_end_str = config['start_date']
    
    print(f"\n🔥 获取预热数据...")
    print(f"预热时间范围: {warmup_start_str} 到 {warmup_end_str}")
    
    warmup_df = db_service.get_kline_data(config['long_coin'], warmup_start_str, warmup_end_str)
    
    if not warmup_df.empty:
        print(f"📈 预热数据: {len(warmup_df)} 个数据点")
        
        # 准备预热数据（包含完整的OHLC和时间戳）
        warmup_data = []
        for _, row in warmup_df.iterrows():
            warmup_data.append({
                'timestamp': row['timestamp'],
                'open': row['open'],
                'high': row['high'],
                'low': row['low'],
                'close': row['close']
            })
        
        # 执行预热
        strategy.warmup_filter(warmup_data)
    else:
        print("⚠️  未获取到预热数据，将使用默认初始化")

    print(f"\n🚀 开始正式回测，共 {len(df)} 个数据点")

    # 🎨 初始化图表数据收集
    chart_data = {
        'timeData': [],
        'klineData': [],
        'sarData30m': [],
        'sarData15m': [],
        'maData15m': [],
        'tradeSignals': [],
        'config_info': {
            'start_date': config['start_date'],
            'end_date': config['end_date'],
            'timeframe': config['timeframe'],
            'sar_start': config['sar_start'],
            'sar_increment': config['sar_increment'],
            'sar_maximum': config['sar_maximum'],
            'initial_capital': config['initial_capital']
        }
    }
    
    # K线聚合状态（动态根据配置调整）
    current_kline = None
    current_kline_start = None

    # 执行策略
    for index, row in df.iterrows():
        timestamp = row['timestamp']
        open_price = row['open']
        high_price = row['high']
        low_price = row['low']
        close_price = row['close']

        # 🎨 收集K线数据（用于图表，根据配置动态调整）
        timeframe = config['timeframe']
        
        # 根据时间周期计算K线开始时间
        if timeframe == '15m':
            minute = timestamp.minute
            if minute < 15:
                period_minute = 0
            elif minute < 30:
                period_minute = 15
            elif minute < 45:
                period_minute = 30
            else:
                period_minute = 45
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        elif timeframe == '20m':
            minute = timestamp.minute
            if minute < 20:
                period_minute = 0
            elif minute < 40:
                period_minute = 20
            else:
                period_minute = 40
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        elif timeframe == '30m':
            minute = timestamp.minute
            if minute < 30:
                period_minute = 0
            else:
                period_minute = 30
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        elif timeframe == '1h':
            kline_start = timestamp.replace(minute=0, second=0, microsecond=0)
        elif timeframe == '2h':
            hour = timestamp.hour
            period_hour = (hour // 2) * 2
            kline_start = timestamp.replace(hour=period_hour, minute=0, second=0, microsecond=0)
        elif timeframe == '4h':
            hour = timestamp.hour
            period_hour = (hour // 4) * 4
            kline_start = timestamp.replace(hour=period_hour, minute=0, second=0, microsecond=0)
        else:
            # 默认30分钟
            minute = timestamp.minute
            if minute < 30:
                period_minute = 0
            else:
                period_minute = 30
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        
        # 检查是否是新的K线周期
        if current_kline_start is None or kline_start != current_kline_start:
            # 保存上一个K线和指标数据
            if current_kline is not None:
                chart_data['timeData'].append(current_kline_start.strftime('%Y-%m-%d %H:%M:%S'))
                chart_data['klineData'].append([
                    round(current_kline['open'], 2),
                    round(current_kline['close'], 2),
                    round(current_kline['low'], 2),
                    round(current_kline['high'], 2)
                ])
                
                # 收集上一个K线周期结束时的SAR值
                current_sar = strategy.sar_indicator.get_stop_loss_level()
                if current_sar is not None:
                    chart_data['sarData30m'].append(round(current_sar, 2))
                    chart_data['sarData15m'].append(round(current_sar, 2))
                    chart_data['maData15m'].append(round(current_sar, 2))
            
            # 开始新的K线周期
            current_kline_start = kline_start
            current_kline = {
                'open': open_price,
                'high': high_price,
                'low': low_price,
                'close': close_price
            }
        else:
            # 更新当前K线
            current_kline['high'] = max(current_kline['high'], high_price)
            current_kline['low'] = min(current_kline['low'], low_price)
            current_kline['close'] = close_price

        # 更新趋势滤波器策略
        result = strategy.update(timestamp, open_price, high_price, low_price, close_price)
        
        # 处理交易信号
        for signal in result['signals']:
            # 🔧 使用更精确的时间戳：如果有exit_timestamp且不为None就用它，否则用当前时间戳
            signal_timestamp = signal.get('exit_timestamp') if signal.get('exit_timestamp') is not None else timestamp
            
            trade_info = {
                'timestamp': signal_timestamp,
                'signal_type': signal['type'],
                'price': signal['price'],
                'reason': signal['reason']
            }
            
            # 添加止损止盈信息（只有开仓信号才有）
            if 'stop_loss' in signal:
                trade_info['stop_loss'] = signal['stop_loss']
            if 'take_profit' in signal:
                trade_info['take_profit'] = signal['take_profit']
            
            # 🔄 添加复投相关信息
            if 'invested_amount' in signal:
                trade_info['invested_amount'] = signal['invested_amount']
            if 'position_shares' in signal:
                trade_info['position_shares'] = signal['position_shares']
            if 'exit_amount' in signal:
                trade_info['exit_amount'] = signal['exit_amount']
            if 'cash_balance' in signal:
                trade_info['cash_balance'] = signal['cash_balance']
            if 'old_balance' in signal:
                trade_info['old_balance'] = signal['old_balance']
            if 'new_balance' in signal:
                trade_info['new_balance'] = signal['new_balance']
            
            if 'profit_loss' in signal:
                trade_info['profit_loss'] = signal['profit_loss']

            if 'transaction_fee' in signal:
                trade_info['transaction_fee'] = signal['transaction_fee']
            
            trades.append(trade_info)
            
            # 🎨 收集交易信号（用于图表显示）
            chart_signal = {
                'time': signal_timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'type': signal['type'],
                'price': round(signal['price'], 2),
                'reason': signal['reason']
            }
            chart_data['tradeSignals'].append(chart_signal)
            
            # 🔍 调试信息：验证复投数据传递
            if signal['type'] in ['OPEN_LONG', 'OPEN_SHORT']:
                invested = trade_info.get('invested_amount', 'N/A')
                balance = trade_info.get('cash_balance', 'N/A')
                print(f"    🔄 开仓数据验证: 投入金额={invested} | 现金余额={balance}")
            elif signal['type'] in ['TAKE_PROFIT_LONG', 'TAKE_PROFIT_SHORT', 'STOP_LOSS_LONG', 'STOP_LOSS_SHORT']:
                old_balance = trade_info.get('old_balance', 'N/A')
                new_balance = trade_info.get('new_balance', 'N/A')
                print(f"    🔄 平仓数据验证: 旧余额={old_balance} | 新余额={new_balance}")
            
            # 更新资金状态和详细信息
            if signal['type'] == 'OPEN_LONG':
                current_position = 'long'
                stop_loss = signal.get('stop_loss', 0)
                take_profit = signal.get('take_profit', None)
                risk = abs(signal['price'] - stop_loss) if stop_loss else 0
                
                if take_profit is not None:
                    reward = abs(take_profit - signal['price'])
                    risk_reward_ratio = reward / risk if risk > 0 else 0
                    print(f"    💰 资金: ${initial_capital:,.2f} | 风险: ${risk:.2f} | 预期收益: ${reward:.2f} | 风险收益比: 1:{risk_reward_ratio:.2f}")
                else:
                    print(f"    💰 资金: ${initial_capital:,.2f} | 风险: ${risk:.2f} | 双周期动态止损策略 (无固定止盈目标)")
                    
            elif signal['type'] == 'OPEN_SHORT':
                current_position = 'short'
                stop_loss = signal.get('stop_loss', 0)
                take_profit = signal.get('take_profit', None)
                risk = abs(stop_loss - signal['price']) if stop_loss else 0
                
                if take_profit is not None:
                    reward = abs(signal['price'] - take_profit)
                    risk_reward_ratio = reward / risk if risk > 0 else 0
                    print(f"    💰 资金: ${initial_capital:,.2f} | 风险: ${risk:.2f} | 预期收益: ${reward:.2f} | 风险收益比: 1:{risk_reward_ratio:.2f}")
                else:
                    print(f"    💰 资金: ${initial_capital:,.2f} | 风险: ${risk:.2f} | 双周期动态止损策略 (无固定止盈目标)")
            elif signal['type'] in ['STOP_LOSS_LONG', 'STOP_LOSS_SHORT', 'TAKE_PROFIT_LONG', 'TAKE_PROFIT_SHORT']:
                current_position = None

        # 每10000个数据点保存一次状态
        if index % 10000 == 0:
            print(f"已处理 {index + 1}/{len(df)} 个数据点")

    # 🎨 保存最后一个K线
    if current_kline is not None:
        chart_data['timeData'].append(current_kline_start.strftime('%Y-%m-%d %H:%M:%S'))
        chart_data['klineData'].append([
            round(current_kline['open'], 2),
            round(current_kline['close'], 2),
            round(current_kline['low'], 2),
            round(current_kline['high'], 2)
        ])
        
        # 收集最后的SAR数据
        final_sar = strategy.sar_indicator.get_stop_loss_level()
        if final_sar is not None:
            chart_data['sarData30m'].append(round(final_sar, 2))
            chart_data['sarData15m'].append(round(final_sar, 2))
            chart_data['maData15m'].append(round(final_sar, 2))

    # === 回测结果统计 ===
    print(f"\n" + "=" * 60)
    print("📊 SAR策略回测结果")
    print("=" * 60)
    
    total_trades = len(trades)
    if total_trades > 0:
        # 统计交易类型
        long_opens = len([t for t in trades if t['signal_type'] == 'OPEN_LONG'])
        short_opens = len([t for t in trades if t['signal_type'] == 'OPEN_SHORT'])
        long_stops = len([t for t in trades if t['signal_type'] == 'STOP_LOSS_LONG'])
        short_stops = len([t for t in trades if t['signal_type'] == 'STOP_LOSS_SHORT'])
        long_profits = len([t for t in trades if t['signal_type'] == 'TAKE_PROFIT_LONG'])
        short_profits = len([t for t in trades if t['signal_type'] == 'TAKE_PROFIT_SHORT'])
        
        # 计算盈亏
        total_pnl = sum([t.get('profit_loss', 0) for t in trades if 'profit_loss' in t])
        
        print(f"📈 总交易次数: {total_trades}")
        print(f"🟢 开多次数: {long_opens}")
        print(f"🔴 开空次数: {short_opens}")
        print(f"✅ 多头止盈: {long_profits}")
        print(f"✅ 空头止盈: {short_profits}")
        print(f"❌ 多头止损: {long_stops}")
        print(f"❌ 空头止损: {short_stops}")
        print(f"💰 总盈亏: ${total_pnl:,.2f}")
        
        # 止盈止损比例
        total_closes = long_stops + short_stops + long_profits + short_profits
        if total_closes > 0:
            profit_rate = (long_profits + short_profits) / total_closes * 100
            loss_rate = (long_stops + short_stops) / total_closes * 100
            print(f"📊 止盈率: {profit_rate:.1f}% | 止损率: {loss_rate:.1f}%")
        
        # 显示最近几笔交易
        print(f"\n📋 最近5笔交易:")
        for trade in trades[-5:]:
            timestamp_str = trade['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if trade.get('timestamp') else 'N/A'
            pnl_str = f", 盈亏: ${trade['profit_loss']:,.2f}" if 'profit_loss' in trade else ""
            print(f"  {timestamp_str} | {trade['signal_type']} | 价格: ${trade['price']:,.2f}{pnl_str}")
            print(f"    原因: {trade['reason']}")
    else:
        print("❌ 未产生任何交易信号")
    
    # 显示最终策略状态
    final_status = strategy.get_current_status()
    print(f"\n🎯 最终策略状态:")
    print(f"  持仓状态: {final_status['position'] or '空仓'}")
    if final_status['entry_price']:
        print(f"  入场价格: ${final_status['entry_price']:,.2f}")
    if final_status['stop_loss_level']:
        print(f"  止损位: ${final_status['stop_loss_level']:,.2f}")
    if final_status['take_profit_level']:
        print(f"  止盈位: ${final_status['take_profit_level']:,.2f}")
    if final_status['sar_value']:
        print(f"  SAR值: ${final_status['sar_value']:,.2f}")
    
    print(f"\n💰 资金状态:")
    print(f"  初始资金: ${initial_capital:,.2f} (单账户模式)")
    
    # === 绩效分析 ===
    print(f"\n" + "=" * 60)
    print("📈 绩效分析")
    print("=" * 60)
    
    # 创建绩效分析器
    analyzer = PerformanceAnalyzer(config)
    
    # 准备价格数据 (timestamp, open, high, low, close)
    price_data = []
    for _, row in df.iterrows():
        price_data.append((row['timestamp'], row['open'], row['high'], row['low'], row['close']))
    
    # 计算每日净值
    nav_df = analyzer.calculate_daily_nav(trades, price_data)
    
    if not nav_df.empty:
        # 计算回撤
        nav_df = analyzer.calculate_drawdown(nav_df)
        
        # 计算关键绩效指标
        metrics = analyzer.calculate_performance_metrics(nav_df, trades)
        
        if metrics:
            # 创建统一的输出目录
            output_dir = create_output_directory(config, metrics['annualized_return'])
            
            print(f"\n📊 关键绩效指标:")
            print(f"  💰 总收益率: {metrics['total_return']:+.2f}%")
            print(f"  📈 年化收益率: {metrics['annualized_return']:+.2f}%")
            print(f"  📉 年化波动率: {metrics['volatility']:.2f}%")
            print(f"  🎯 夏普比率: {metrics['sharpe_ratio']:.2f}")
            print(f"  ⚠️  最大回撤: {metrics['max_drawdown']:.2f}%")
            print(f"  🎯 胜率: {metrics['win_rate']:.1f}%")
            print(f"  📊 盈亏比: {metrics['profit_loss_ratio']:.2f}")
            print(f"  🔄 总交易次数: {metrics['total_trades']}")
            print(f"  📅 交易天数: {metrics['trading_days']}")
            print(f"  💎 最终净值: ${metrics['final_nav']:,.2f}")
            
            # === 生成所有结果文件到统一目录 ===
            print(f"\n" + "=" * 60)
            print("📁 生成结果文件")
            print("=" * 60)
            
            # 1. 导出交易记录到Excel
            try:
                excel_file = export_trades_to_excel(trades, config, output_dir)
                if excel_file:
                    print(f"📋 交易记录: {os.path.abspath(excel_file)}")
            except Exception as e:
                print(f"⚠️  交易记录导出失败: {e}")
            
            # 2. 生成绩效图表
            try:
                chart_files = analyzer.generate_performance_charts(nav_df, output_dir)
                print(f"📈 绩效图表已生成")
            except Exception as e:
                print(f"⚠️  绩效图表生成失败: {e}")
            
            # 3. 生成绩效Excel报告
            try:
                perf_excel = analyzer.generate_performance_excel(nav_df, output_dir)
                if perf_excel:
                    print(f"📋 绩效Excel报告: {os.path.abspath(perf_excel)}")
            except Exception as e:
                print(f"⚠️  绩效Excel报告生成失败: {e}")
            
            # 4. 生成HTML报告
            try:
                html_report = analyzer.generate_html_report(nav_df, config['long_coin'], trades, config, output_dir)
                if html_report:
                    print(f"🌐 HTML绩效报告: {os.path.abspath(html_report)}")
            except Exception as e:
                print(f"⚠️  HTML报告生成失败: {e}")
            
            # 5. 生成交互式图表
            try:
                chart_html = create_chart_html(chart_data, output_dir)
                if chart_html:
                    print(f"🎨 交互式图表: {os.path.abspath(chart_html)}")
            except Exception as e:
                print(f"⚠️  交互式图表生成失败: {e}")
            
            print(f"\n✅ 所有结果文件已保存到: {os.path.abspath(output_dir)}")
    else:
        print("⚠️  无法计算净值，跳过绩效分析")
    
    # 关闭数据库连接
    db_service.disconnect()
    
    print(f"\n🎉 回测完成！所有结果文件已按币种和年化收益率分类保存")

if __name__ == "__main__":
    main() 