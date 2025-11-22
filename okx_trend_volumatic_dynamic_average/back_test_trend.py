#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# æ·»åŠ å½“å‰ç›®å½•åˆ°Pythonè·¯å¾„
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from trend_volumatic_dynamic_average_strategy import TrendVolumaticDynamicAverageStrategy
from database_config import LOCAL_DATABASE_CONFIG
from database_service import DatabaseService
from strategy_configs import get_strategy_config, print_config_info
from performance_analysis import PerformanceAnalyzer

def create_output_directory(config, annual_return):
    """
    åˆ›å»ºç»Ÿä¸€çš„è¾“å‡ºç›®å½•ç»“æ„
    /back_test_data/{å¸ç§}/{å¹´åŒ–æ”¶ç›Šç‡}-{æ—¶é—´æˆ³}/
    
    Args:
        config: ç­–ç•¥é…ç½®
        annual_return: å¹´åŒ–æ”¶ç›Šç‡ï¼ˆç”¨äºæ–‡ä»¶å¤¹å‘½åï¼‰
    
    Returns:
        str: è¾“å‡ºç›®å½•è·¯å¾„
    """
    # è·å–å¸ç§
    long_coin = config['long_coin']
    
    # ç”Ÿæˆæ—¶é—´æˆ³
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # æ ¼å¼åŒ–å¹´åŒ–æ”¶ç›Šç‡ï¼ˆä¿ç•™2ä½å°æ•°ï¼‰
    annual_return_str = f"{annual_return:+.2f}%"
    
    # åˆ›å»ºç›®å½•ç»“æ„
    base_dir = "back_test_data"
    coin_dir = os.path.join(base_dir, long_coin)
    result_dir = os.path.join(coin_dir, f"{annual_return_str}-{timestamp}")
    
    # ç¡®ä¿ç›®å½•å­˜åœ¨
    os.makedirs(result_dir, exist_ok=True)
    
    print(f"ğŸ“ åˆ›å»ºè¾“å‡ºç›®å½•: {result_dir}")
    return result_dir

def export_trades_to_excel(trades, config, output_dir):
    """
    å¯¼å‡ºäº¤æ˜“è®°å½•åˆ°Excelæ–‡ä»¶
    
    Args:
        trades: äº¤æ˜“è®°å½•åˆ—è¡¨
        config: ç­–ç•¥é…ç½®
        output_dir: è¾“å‡ºç›®å½•
    """
    if not trades:
        print("âŒ æ²¡æœ‰äº¤æ˜“è®°å½•ï¼Œè·³è¿‡Excelå¯¼å‡º")
        return None
    
    # ç”Ÿæˆæ–‡ä»¶åï¼ˆæ·»åŠ æ—¶é—´æˆ³ï¼‰
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(output_dir, f"äº¤æ˜“è®°å½•_{timestamp}.xlsx")
    
    # åˆå¹¶å¼€ä»“å’Œå¹³ä»“è®°å½•ï¼ŒæŒ‰æ—¶é—´é¡ºåº
    trade_data = []
    current_entry = None  # è®°å½•å½“å‰å¼€ä»“ä¿¡æ¯
    
    for i, trade in enumerate(trades):
        if trade['signal_type'] in ['OPEN_LONG', 'OPEN_SHORT']:
            # å¼€ä»“è®°å½•ï¼ˆæ»¡ä»“æ¨¡å¼ï¼‰
            if 'LONG' in trade['signal_type']:
                direction = "åšå¤š"
            else:
                direction = "åšç©º"
            
            operation_type = "å¼€ä»“"
            stop_loss = trade.get('stop_loss', 'N/A')
            take_profit = trade.get('take_profit', 'N/A')
            
            # ğŸ”„ è·å–å®é™…æŠ•å…¥é‡‘é¢ï¼ˆæ¥è‡ªç­–ç•¥ä¼ é€’çš„å¤æŠ•ä¿¡æ¯ï¼‰
            trade_amount = trade.get('invested_amount', 0)
            
            # è°ƒè¯•ä¿¡æ¯ï¼šæ£€æŸ¥å¼€ä»“ä¿¡å·æ˜¯å¦åŒ…å«æŠ•å…¥é‡‘é¢
            if trade_amount == 0:
                print(f"âš ï¸  è­¦å‘Šï¼šå¼€ä»“ä¿¡å· {trade['signal_type']} ç¼ºå°‘ invested_amount ä¿¡æ¯")
                print(f"    ä¿¡å·å†…å®¹: {trade}")
                # ğŸ”„ ä½¿ç”¨å½“å‰ç°é‡‘ä½™é¢è®¡ç®—ï¼ˆå¤æŠ•é€»è¾‘ï¼‰
                position_size = config.get('position_size_percentage', 100) / 100
                # å•è´¦æˆ·æ¨¡å¼ï¼šä½¿ç”¨ç»Ÿä¸€ç°é‡‘ä½™é¢
                current_balance = trade.get('cash_balance', config['initial_capital'])
                trade_amount = current_balance * position_size
                print(f"    ğŸ”„ å¤æŠ•è®¡ç®—: å½“å‰ä½™é¢${current_balance:,.2f} Ã— {position_size*100}% = ${trade_amount:,.2f}")
            
            # å•è´¦æˆ·æ¨¡å¼ï¼šè·å–ç»Ÿä¸€ç°é‡‘ä½™é¢ä¿¡æ¯
            cash_balance = trade.get('cash_balance', 0)
            
            # ä¿å­˜å½“å‰å¼€ä»“ä¿¡æ¯ï¼Œç”¨äºå¹³ä»“æ—¶å¼•ç”¨
            current_entry = {
                'direction': direction,
                'price': trade['price'],
                'stop_loss': stop_loss,
                'take_profit': take_profit,
                'amount': trade_amount,
                'cash_balance': cash_balance
            }
            
            # è·å–ä»½é¢ä¿¡æ¯ï¼ˆæ»¡ä»“æ¨¡å¼ï¼‰
            current_shares = trade.get('position_shares', 0)
            
            trade_data.append({
                'åºå·': len(trade_data) + 1,
                'æ“ä½œç±»å‹': operation_type,
                'äº¤æ˜“æ–¹å‘': direction,
                'æ—¶é—´': trade['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if trade.get('timestamp') else 'N/A',
                'ä»·æ ¼': f"{trade['price']:.2f}",
                'äº¤æ˜“é‡‘é¢': f"{trade_amount:,.2f}",
                'äº¤æ˜“ä»½é¢': f"{current_shares:.4f}",
                'æ­¢æŸä½': f"{stop_loss:.2f}" if isinstance(stop_loss, (int, float)) else stop_loss,
                'æ­¢ç›ˆä½': f"{take_profit:.2f}" if isinstance(take_profit, (int, float)) else take_profit,
                'ç›ˆäºé‡‘é¢': '-',
                'æ‰‹ç»­è´¹': f"{trade.get('transaction_fee', 0):.2f}",
                'æ”¶ç›Šç‡': '-',
                'äº¤æ˜“ç»“æœ': '-',
                'ç°é‡‘ä½™é¢': f"{cash_balance:,.2f}",
                'åŸå› ': trade['reason']
            })
            
        elif trade['signal_type'] in ['STOP_LOSS_LONG', 'STOP_LOSS_SHORT', 'TAKE_PROFIT_LONG', 'TAKE_PROFIT_SHORT', 
                                       'MA_PROFIT_LONG', 'MA_LOSS_LONG', 'MA_PROFIT_SHORT', 'MA_LOSS_SHORT',
                                       'MAX_STOP_LOSS_LONG', 'MAX_STOP_LOSS_SHORT']:
            # å¹³ä»“è®°å½•
            direction = "åšå¤š" if 'LONG' in trade['signal_type'] else "åšç©º"
            profit_loss = trade.get('profit_loss', 0)
            
            # åˆ¤æ–­å¹³ä»“ç±»å‹å’Œç»“æœ - æ›´ç²¾ç¡®åœ°æè¿°å¹³ä»“åŸå› 
            if 'TAKE_PROFIT' in trade['signal_type']:
                close_type = "å›ºå®šæ­¢ç›ˆ"
                # ğŸ”´ æ ¹æ®å®é™…ç›ˆäºåˆ¤æ–­ç»“æœ
                result = "ç›ˆåˆ©" if profit_loss > 0 else "äºæŸ"
                print(f"    ğŸ¯ å¹³ä»“åˆ†ç±»: {trade['signal_type']} + {'ç›ˆåˆ©' if profit_loss > 0 else 'äºæŸ'}${profit_loss:.2f} -> {close_type} ({result})")
            elif 'MA_PROFIT' in trade['signal_type'] or 'MA_LOSS' in trade['signal_type']:
                # å›å½’MAå–å‡ºä¿¡å·
                if profit_loss > 0:
                    close_type = "å›å½’MAç›ˆåˆ©"
                    result = "ç›ˆåˆ©"
                    print(f"    ğŸ“Š å¹³ä»“åˆ†ç±»: {trade['signal_type']} + ç›ˆåˆ©${profit_loss:.2f} -> {close_type} ({result})")
                else:
                    close_type = "å›å½’MAäºæŸ"
                    result = "äºæŸ"
                    print(f"    ğŸ“Š å¹³ä»“åˆ†ç±»: {trade['signal_type']} + äºæŸ${profit_loss:.2f} -> {close_type} ({result})")
            elif 'MAX_STOP_LOSS' in trade['signal_type']:
                # æœ€å¤§å›ºå®šæ­¢æŸ
                close_type = "æœ€å¤§å›ºå®šæ­¢æŸ"
                result = "äºæŸ"
                print(f"    ğŸ”’ å¹³ä»“åˆ†ç±»: {trade['signal_type']} + äºæŸ${profit_loss:.2f} -> {close_type} ({result})")
            else:  # STOP_LOSS - åŸºäºå›ºå®šæ¯”ä¾‹æˆ–å¸ƒæ—å¸¦å®½åº¦çš„æ­¢æŸ
                if profit_loss > 0:
                    close_type = "åŠ¨æ€æ­¢ç›ˆ"  # å…¶ä»–åŸå› å¯¼è‡´çš„ç›ˆåˆ©å¹³ä»“
                    result = "ç›ˆåˆ©"
                    print(f"    ğŸ“ˆ å¹³ä»“åˆ†ç±»: {trade['signal_type']} + ç›ˆåˆ©${profit_loss:.2f} -> {close_type} ({result})")
                else:
                    close_type = "åŠ¨æ€æ­¢æŸ"  # å›ºå®šæ¯”ä¾‹æˆ–å¸ƒæ—å¸¦æ­¢æŸ
                    result = "äºæŸ"
                    print(f"    ğŸ“‰ å¹³ä»“åˆ†ç±»: {trade['signal_type']} + äºæŸ${profit_loss:.2f} -> {close_type} ({result})")
            
            # è®¡ç®—æ”¶ç›Šç‡ï¼ˆåŸºäºæŠ•å…¥æœ¬é‡‘ï¼‰
            if profit_loss != 0 and current_entry and current_entry.get('amount', 0) > 0:
                # ä½¿ç”¨æŠ•å…¥çš„èµ„é‡‘ä½œä¸ºåŸºå‡†è®¡ç®—æ”¶ç›Šç‡
                invested_capital = current_entry['amount']
                return_rate = (profit_loss / invested_capital) * 100
            else:
                return_rate = 0
            
            # ä½¿ç”¨å¼€ä»“æ—¶çš„æ­¢æŸæ­¢ç›ˆä½å’Œæœ€æ–°ç°é‡‘ä½™é¢
            if current_entry:
                entry_stop_loss = current_entry['stop_loss']
                entry_take_profit = current_entry['take_profit']
                entry_amount = current_entry['amount']
            else:
                # å¦‚æœæ²¡æœ‰å¼€ä»“è®°å½•ï¼Œä½¿ç”¨é»˜è®¤å€¼
                print(f"âš ï¸  è­¦å‘Šï¼šå¹³ä»“è®°å½• {trade['signal_type']} æ²¡æœ‰å¯¹åº”çš„å¼€ä»“è®°å½•")
                entry_stop_loss = 'N/A'
                entry_take_profit = 'N/A'
                entry_amount = trade.get('invested_amount', 0)
                
            # ç¡®ä¿ entry_amount ä¸ä¸º0ï¼Œé¿å…é™¤é›¶é”™è¯¯
            if entry_amount == 0:
                print(f"âš ï¸  è­¦å‘Šï¼šäº¤æ˜“é‡‘é¢ä¸º0ï¼Œè®¾ç½®ä¸ºé»˜è®¤å€¼1000")
                entry_amount = 1000  # è®¾ç½®ä¸€ä¸ªé»˜è®¤å€¼é¿å…é™¤é›¶é”™è¯¯
            
            # ğŸ”„ è·å–å¹³ä»“åçš„ç°é‡‘ä½™é¢ï¼ˆå•è´¦æˆ·æ¨¡å¼ï¼‰
            new_cash_balance = trade.get('new_balance', 0)
            
            # ğŸ”„ å¹³ä»“äº¤æ˜“é‡‘é¢ = æœ¬é‡‘ + ç›ˆäº
            close_trade_amount = entry_amount + profit_loss
            print(f"    ğŸ’° å¹³ä»“é‡‘é¢è®¡ç®—: æœ¬é‡‘${entry_amount:,.2f} + ç›ˆäº${profit_loss:+.2f} = ${close_trade_amount:,.2f}")
            
            trade_data.append({
                'åºå·': len(trade_data) + 1,
                'æ“ä½œç±»å‹': f'å¹³ä»“({close_type})',
                'äº¤æ˜“æ–¹å‘': direction,
                'æ—¶é—´': trade['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if trade.get('timestamp') else 'N/A',
                'ä»·æ ¼': f"{trade['price']:.2f}",
                'äº¤æ˜“é‡‘é¢': f"{close_trade_amount:,.2f}",
                'äº¤æ˜“ä»½é¢': f"{trade.get('position_shares', 0):.4f}",
                'æ­¢æŸä½': f"{entry_stop_loss:.2f}" if isinstance(entry_stop_loss, (int, float)) else entry_stop_loss,
                'æ­¢ç›ˆä½': f"{entry_take_profit:.2f}" if isinstance(entry_take_profit, (int, float)) else entry_take_profit,
                'ç›ˆäºé‡‘é¢': f"{profit_loss:.2f}",
                'æ‰‹ç»­è´¹': f"{trade.get('transaction_fee', 0):.2f}",
                'æ”¶ç›Šç‡': f"{return_rate:+.2f}%",
                'äº¤æ˜“ç»“æœ': result,
                'ç°é‡‘ä½™é¢': f"{new_cash_balance:,.2f}",
                'åŸå› ': trade['reason']
            })
            
            # æ¸…é™¤å½“å‰å¼€ä»“ä¿¡æ¯
            current_entry = None
    
    # åˆ›å»ºDataFrame
    df = pd.DataFrame(trade_data)
    
    # å†™å…¥Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # äº¤æ˜“æ˜ç»†è¡¨
        df.to_excel(writer, sheet_name='äº¤æ˜“æ˜ç»†', index=False)
        
        # è·å–å·¥ä½œç°¿å’Œå·¥ä½œè¡¨
        workbook = writer.book
        worksheet = writer.sheets['äº¤æ˜“æ˜ç»†']
        
        # è®¾ç½®æ ·å¼
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # è®¾ç½®è¡¨å¤´æ ·å¼
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # è®¾ç½®æ•°æ®è¡Œæ ·å¼
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = border
                
                # æ ¹æ®æ“ä½œç±»å‹å’Œäº¤æ˜“ç»“æœè®¾ç½®é¢œè‰²
                if cell.column == 2:  # æ“ä½œç±»å‹åˆ—
                    if cell.value == "å¼€ä»“":
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        cell.font = Font(color="0066CC")
                    elif "å¹³ä»“" in str(cell.value):
                        if "æ­¢ç›ˆ" in str(cell.value):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif "æ­¢æŸ" in str(cell.value):
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                elif cell.column == 6:  # äº¤æ˜“é‡‘é¢åˆ—
                    cell.font = Font(bold=True)
                    if "å¼€ä»“" in worksheet.cell(row=cell.row, column=2).value:
                        cell.font = Font(color="0066CC", bold=True)
                elif cell.column == 11:  # äº¤æ˜“ç»“æœåˆ—ï¼ˆä½ç½®å˜äº†ï¼Œå› ä¸ºåŠ äº†ç°é‡‘ä½™é¢åˆ—ï¼‰
                    if cell.value == "ç›ˆåˆ©":
                        cell.font = Font(color="006100", bold=True)
                    elif cell.value == "äºæŸ":
                        cell.font = Font(color="9C0006", bold=True)
                elif cell.column == 12:  # ç°é‡‘ä½™é¢åˆ—
                    cell.font = Font(bold=True, color="FF6600")  # æ©™è‰²åŠ ç²—æ˜¾ç¤ºç°é‡‘ä½™é¢
        
        # è‡ªåŠ¨è°ƒæ•´åˆ—å®½
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # æ·»åŠ ç»Ÿè®¡ä¿¡æ¯è¡¨
        stats_data = []
        if trade_data:
            # åˆ†ç¦»å¼€ä»“å’Œå¹³ä»“æ“ä½œè¿›è¡Œç»Ÿè®¡
            open_trades = [t for t in trade_data if t['æ“ä½œç±»å‹'] == 'å¼€ä»“']
            close_trades = [t for t in trade_data if 'å¹³ä»“' in t['æ“ä½œç±»å‹']]
            
            total_complete_trades = len(close_trades)  # å®Œæ•´äº¤æ˜“æ¬¡æ•°
            profit_trades = len([t for t in close_trades if t['äº¤æ˜“ç»“æœ'] == 'ç›ˆåˆ©'])
            loss_trades = len([t for t in close_trades if t['äº¤æ˜“ç»“æœ'] == 'äºæŸ'])
            win_rate = (profit_trades / total_complete_trades) * 100 if total_complete_trades > 0 else 0
            
            # è®¡ç®—æ€»ç›ˆäºï¼ˆåªè®¡ç®—å¹³ä»“è®°å½•ï¼‰
            total_pnl = sum([float(t['ç›ˆäºé‡‘é¢'].replace('$', '')) for t in close_trades if t['ç›ˆäºé‡‘é¢'] != '-'])
            
            # ç»Ÿè®¡å¼€ä»“ç±»å‹ï¼ˆæ»¡ä»“æ¨¡å¼ï¼Œæ— åŠ ä»“ï¼‰
            long_opens = len([t for t in open_trades if t['äº¤æ˜“æ–¹å‘'] == 'åšå¤š'])
            short_opens = len([t for t in open_trades if t['äº¤æ˜“æ–¹å‘'] == 'åšç©º'])
            
            # ç»Ÿè®¡å¹³ä»“ç±»å‹
            take_profit_closes = len([t for t in close_trades if 'æ­¢ç›ˆ' in t['æ“ä½œç±»å‹']])
            stop_loss_closes = len([t for t in close_trades if 'æ­¢æŸ' in t['æ“ä½œç±»å‹']])
            ma_closes = len([t for t in close_trades if 'å›å½’MA' in t['æ“ä½œç±»å‹']])
            
            stats_data = [
                ['äº¤æ˜“ç»Ÿè®¡', ''],
                ['æ€»å¼€ä»“æ¬¡æ•°', len(open_trades)],
                ['æ€»å¹³ä»“æ¬¡æ•°', len(close_trades)],
                ['å®Œæ•´äº¤æ˜“æ¬¡æ•°', total_complete_trades],
                ['åšå¤šå¼€ä»“', long_opens],
                ['åšç©ºå¼€ä»“', short_opens],
                ['æ­¢ç›ˆå¹³ä»“', take_profit_closes],
                ['æ­¢æŸå¹³ä»“', stop_loss_closes],
                ['å›å½’MAå¹³ä»“', ma_closes],
                ['ç›ˆåˆ©æ¬¡æ•°', profit_trades],
                ['äºæŸæ¬¡æ•°', loss_trades],
                ['èƒœç‡', f"{win_rate:.2f}%"],
                ['æ€»ç›ˆäº', f"${total_pnl:.2f}"],
                ['', ''],
                ['ç­–ç•¥å‚æ•°', ''],
                ['å•å‘¨æœŸæ¨¡å¼', config['timeframe']],
                ['VIDYAé•¿åº¦', config.get('vidya_length', 20)],
                ['VIDYAåŠ¨é‡', config.get('vidya_momentum', 9)],
                ['VIDYAå¹³æ»‘', config.get('vidya_smooth', 15)],
                ['ATRè·ç¦»', config.get('vidya_band_distance', 2.0)],
                ['ATRå‘¨æœŸ', config.get('vidya_atr_period', 200)],
                ['å›ºå®šæ­¢ç›ˆ', f"{config['fixed_take_profit_pct']}%"],
                ['æ­¢æŸç­–ç•¥', 'VIDYAåŠ¨æ€æ­¢æŸè·Ÿéš'],
                ['å¼€ä»“æœºåˆ¶', 'VIDYAè¶‹åŠ¿æ”¹å˜æ—¶å¼€ä»“']
            ]
        
        stats_df = pd.DataFrame(stats_data, columns=['é¡¹ç›®', 'æ•°å€¼'])
        stats_df.to_excel(writer, sheet_name='ç»Ÿè®¡æ±‡æ€»', index=False)
        
        # è®¾ç½®ç»Ÿè®¡è¡¨æ ·å¼
        stats_sheet = writer.sheets['ç»Ÿè®¡æ±‡æ€»']
        for cell in stats_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        for column in stats_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            stats_sheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"âœ… äº¤æ˜“è®°å½•å·²å¯¼å‡ºåˆ°: {filename}")
    open_count = len([t for t in trade_data if t['æ“ä½œç±»å‹'] == 'å¼€ä»“'])
    close_count = len([t for t in trade_data if 'å¹³ä»“' in t['æ“ä½œç±»å‹']])
    print(f"ğŸ“Š å…±å¯¼å‡º {len(trade_data)} æ¡è®°å½• (å¼€ä»“: {open_count}, å¹³ä»“: {close_count})")
    
    return filename

def create_chart_html(chart_data, output_dir):
    """åˆ›å»ºäº¤äº’å¼å›¾è¡¨HTML"""
    import json
    config_info = chart_data.get('config_info', {})
    timeframe = config_info.get('timeframe', '30m')
    
    html_content = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>çº¯VIDYAç­–ç•¥å›æµ‹ - çœŸå®æ•°æ®äº¤äº’å›¾è¡¨</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <style>
        body {{
            margin: 0;
            padding: 20px;
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
            color: #ffffff;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1800px;
            margin: 0 auto;
        }}
        .header {{
            text-align: center;
            margin-bottom: 30px;
            background: rgba(255, 255, 255, 0.08);
            padding: 25px;
            border-radius: 15px;
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .header h1 {{
            margin: 0 0 15px 0;
            color: #00d4ff;
            text-shadow: 0 0 30px rgba(0, 212, 255, 0.4);
            font-size: 2.2em;
        }}
        .config-info {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-top: 20px;
            padding: 20px;
            background: rgba(0, 0, 0, 0.2);
            border-radius: 10px;
        }}
        .config-item {{
            display: flex;
            justify-content: space-between;
            padding: 8px 12px;
            background: rgba(255, 255, 255, 0.05);
            border-radius: 6px;
            border-left: 3px solid #00d4ff;
        }}
        .chart-container {{
            width: 100%;
            height: 800px;
            background: rgba(255, 255, 255, 0.08);
            border-radius: 15px;
            padding: 20px;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.4);
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .legend {{
            display: flex;
            justify-content: center;
            gap: 30px;
            margin-bottom: 25px;
            flex-wrap: wrap;
        }}
        .legend-item {{
            display: flex;
            align-items: center;
            gap: 10px;
            font-size: 15px;
            background: rgba(255, 255, 255, 0.08);
            padding: 10px 15px;
            border-radius: 8px;
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .legend-color {{
            width: 16px;
            height: 16px;
            border-radius: 4px;
            box-shadow: 0 0 10px rgba(255, 255, 255, 0.3);
        }}
        .info-panel {{
            background: rgba(255, 255, 255, 0.08);
            padding: 25px;
            border-radius: 15px;
            margin-top: 30px;
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 15px;
        }}
        .info-row {{
            display: flex;
            justify-content: space-between;
            padding: 12px 15px;
            background: rgba(0, 0, 0, 0.3);
            border-radius: 8px;
            border-left: 4px solid #00d4ff;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>ğŸ¯ çº¯VIDYAç­–ç•¥å›æµ‹ - {timeframe}Kçº¿å›¾è¡¨</h1>
            <p style="color: #cccccc; font-size: 16px;">ğŸ’« VIDYAåŠ¨æ€è¶‹åŠ¿ + å›ºå®šæ­¢ç›ˆ + VIDYAåŠ¨æ€æ­¢æŸ | ğŸ¯ çœŸå®å›æµ‹æ•°æ® | ğŸ“Š äº¤äº’å¼åˆ†æ</p>
            
            <div class="config-info">
                <div class="config-item">
                    <span>ğŸ“… å›æµ‹å¼€å§‹:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('start_date', 'N/A')}</span>
                </div>
                <div class="config-item">
                    <span>ğŸ“… å›æµ‹ç»“æŸ:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('end_date', 'N/A')}</span>
                </div>
                <div class="config-item">
                    <span>ğŸ¯ æ—¶é—´å‘¨æœŸ:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('timeframe', 'N/A')}</span>
                </div>
                <div class="config-item">
                    <span>âš¡ VIDYAå‚æ•°:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{config_info.get('vidya_length', 20)}/{config_info.get('vidya_momentum', 9)}/{config_info.get('vidya_smooth', 15)}</span>
                </div>
                <div class="config-item">
                    <span>ğŸ“ ATRå¸¦å®½:</span>
                    <span style="color: #00d4ff; font-weight: bold;">è·ç¦»{config_info.get('vidya_band_distance', 2.0)} | å‘¨æœŸ{config_info.get('vidya_atr_period', 200)}</span>
                </div>
                <div class="config-item">
                    <span>ğŸ” æ¢è½´ç‚¹:</span>
                    <span style="color: #00d4ff; font-weight: bold;">å·¦{config_info.get('vidya_pivot_left', 3)}/å³{config_info.get('vidya_pivot_right', 3)}</span>
                </div>
                <div class="config-item">
                    <span>ğŸ’° åˆå§‹èµ„é‡‘:</span>
                    <span style="color: #00d4ff; font-weight: bold;">${config_info.get('initial_capital', 0):,}</span>
                </div>
                <div class="config-item">
                    <span>ğŸ“Š Kçº¿æ•°é‡:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{len(chart_data.get('klineData', []))}</span>
                </div>
                <div class="config-item">
                    <span>ğŸ¯ äº¤æ˜“ä¿¡å·:</span>
                    <span style="color: #00d4ff; font-weight: bold;">{len(chart_data.get('tradeSignals', []))}</span>
                </div>
            </div>
        </div>

        <div class="legend">
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #00da3c, #00ff41);"></div>
                <span>ğŸ“ˆ ä¸Šæ¶¨Kçº¿</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #ec0000, #ff1744);"></div>
                <span>ğŸ“‰ ä¸‹è·ŒKçº¿</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #ffaa00, #ffcc00);"></div>
                <span>ğŸ“ VIDYAä¸Šè½¨ (é»„è‰²ç»†çº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #00aaff, #00ccff);"></div>
                <span>ğŸ“ VIDYAä¸‹è½¨ (è“è‰²ç»†çº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #ff6600, #ff8800);"></div>
                <span>ğŸ“ˆ åšå¤šè¶‹åŠ¿çº¿ (æ©™è‰²ç²—çº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #0066ff, #0088ff);"></div>
                <span>ğŸ“‰ åšç©ºè¶‹åŠ¿çº¿ (è“è‰²ç²—çº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #17dfad, #00ff88);"></div>
                <span>ğŸ“ˆ æ”¯æ’‘ä½ (ç»¿è‰²ç»†çº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #dd326b, #ff4081);"></div>
                <span>ğŸ“‰ é˜»åŠ›ä½ (çº¢è‰²ç»†çº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #00ff00, #00dd00);"></div>
                <span>ğŸ“Š EMA50 (ç»¿è‰²çº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #ffff00, #ffdd00);"></div>
                <span>ğŸ“Š EMA120åŸå§‹ (é»„è‰²è™šçº¿)</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: linear-gradient(45deg, #ff00ff, #dd00dd);"></div>
                <span>ğŸ“Š EMA120å¹³æ»‘ (ç´«è‰²çº¿)</span>
            </div>
        </div>

        <div class="chart-container">
            <div id="main-chart" style="width: 100%; height: 100%;"></div>
        </div>

        <div class="info-panel">
            <h3>ğŸ“Š å½“å‰Kçº¿è¯¦æƒ…</h3>
            <div class="info-grid" id="currentInfo">
                <div class="info-row">
                    <span>ğŸ• æ—¶é—´:</span>
                    <span id="currentTime">ç‚¹å‡»Kçº¿æŸ¥çœ‹è¯¦æƒ…</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“ˆ å¼€ç›˜:</span>
                    <span id="currentOpen">-</span>
                </div>
                <div class="info-row">
                    <span>â¬†ï¸ æœ€é«˜:</span>
                    <span id="currentHigh">-</span>
                </div>
                <div class="info-row">
                    <span>â¬‡ï¸ æœ€ä½:</span>
                    <span id="currentLow">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“‰ æ”¶ç›˜:</span>
                    <span id="currentClose">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ¯ VIDYAè¶‹åŠ¿:</span>
                    <span id="currentTrend">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“ VIDYAä¸Šè½¨:</span>
                    <span id="currentUpperBand">-</span>
            </div>
                <div class="info-row">
                    <span>ğŸ“ VIDYAä¸‹è½¨:</span>
                    <span id="currentLowerBand">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ’« CMOåŠ¨é‡:</span>
                    <span id="currentCMO">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“ˆ Buyäº¤æ˜“é‡:</span>
                    <span id="currentBuyVolume">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“‰ Selläº¤æ˜“é‡:</span>
                    <span id="currentSellVolume">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸšï¸ Delta Volume(åŠ¨æ€):</span>
                    <span id="currentDeltaVolume">-</span>
                </div>
                <div class="info-row" style="border-left: 4px solid #00ff00;">
                    <span>âš¡ DVå›ºå®šå‘¨æœŸ(14):</span>
                    <span id="currentDeltaVolumeFixed" style="font-weight: bold;">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“Š EMA50:</span>
                    <span id="currentEma50">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“Š EMA120åŸå§‹:</span>
                    <span id="currentEma120Raw">-</span>
                </div>
                <div class="info-row">
                    <span>ğŸ“Š EMA120å¹³æ»‘:</span>
                    <span id="currentEma120Smoothed">-</span>
                </div>
            </div>
        </div>
    </div>

    <script>
        const backtestData = {json.dumps(chart_data, ensure_ascii=False, indent=8)};
        
        const chartDom = document.getElementById('main-chart');
        const myChart = echarts.init(chartDom, 'dark');
        
        function getChartOption(data) {{
            return {{
                animation: true,
                backgroundColor: 'transparent',
                legend: {{
                    show: true,
                    top: 10,
                    left: 'center',
                    textStyle: {{
                        color: '#ffffff',
                        fontSize: 13
                    }},
                    selectedMode: true,  // ğŸ”´ å¯ç”¨å›¾ä¾‹ç‚¹å‡»æ§åˆ¶
                    selected: {{
                        'Kçº¿': true,
                        'VIDYAä¸Šè½¨': true,
                        'VIDYAä¸‹è½¨': true,
                        'åšå¤šè¶‹åŠ¿çº¿': true,
                        'åšç©ºè¶‹åŠ¿çº¿': true,
                        'æ”¯æ’‘ä½': true,
                        'é˜»åŠ›ä½': true,
                        'EMA50': true,
                        'EMA120åŸå§‹': false,  // é»˜è®¤éšè—åŸå§‹EMA120
                        'EMA120å¹³æ»‘': true
                    }}
                }},
                tooltip: {{
                    trigger: 'axis',
                    axisPointer: {{
                        type: 'cross',
                        lineStyle: {{
                            color: '#00d4ff',
                            width: 1.5,
                            opacity: 0.8
                        }}
                    }},
                    backgroundColor: 'rgba(20, 20, 30, 0.95)',
                    borderColor: '#00d4ff',
                    borderWidth: 2,
                    textStyle: {{
                        color: '#fff',
                        fontSize: 14
                    }},
                    formatter: function (params) {{
                        const dataIndex = params[0].dataIndex;
                        const time = data.timeData[dataIndex];
                        const kline = data.klineData[dataIndex];
                        const volume = data.volumeData[dataIndex];
                        const deltaVolume = data.deltaVolumeData[dataIndex];
                        const trendDirection = data.trendDirection[dataIndex];
                        
                        // ğŸ”´ ä»åç«¯è·å–çœŸå®çš„ç´¯ç§¯Buy/Selläº¤æ˜“é‡ï¼ˆPine Scripté€»è¾‘ï¼‰
                        // åç«¯åœ¨æ•´ä¸ªè¶‹åŠ¿æœŸé—´ç´¯ç§¯ï¼šé˜³çº¿å½’buy_volumeï¼Œé˜´çº¿å½’sell_volume
                        const buyVolume = data.buyVolumeData[dataIndex] || 0;
                        const sellVolume = data.sellVolumeData[dataIndex] || 0;
                        const avgVolume = (buyVolume + sellVolume) / 2;
                        const deltaVolumePercent = avgVolume > 0 ? ((buyVolume - sellVolume) / avgVolume * 100).toFixed(2) : '0.00';
                        
                        // è·å–ä¸Šä¸‹è½¨æ•°æ®
                        const upperBand = data.upperBandData[dataIndex];
                        const lowerBand = data.lowerBandData[dataIndex];
                        
                        // ğŸ”´ è·å–EMAæ•°æ®
                        const ema50 = data.ema50Data[dataIndex];
                        const ema120Raw = data.ema120RawData[dataIndex];
                        const ema120Smoothed = data.ema120SmoothedData[dataIndex];
                        
                        // ğŸ”´ è·å–CMOæ•°æ®
                        const cmo = data.cmoData[dataIndex];
                        
                        // ğŸ”´ è·å–å›ºå®šå‘¨æœŸDelta Volumeç™¾åˆ†æ¯”
                        const deltaVolumePercentFixed = data.deltaVolumePercentFixed[dataIndex];
                        
                        updateInfoPanel(time, kline, null, null, null, null, null, null, null);
                        
                        const change = kline[1] - kline[0];
                        const changePercent = (change / kline[0] * 100).toFixed(3);
                        const changeColor = change >= 0 ? '#00da3c' : '#ec0000';
                        // ğŸ”´ ä¿®æ­£è¶‹åŠ¿é¢œè‰²ï¼šåšå¤šæ—¶çº¢è‰²ï¼Œåšç©ºæ—¶ç»¿è‰²
                        const trendColor = trendDirection === 'up' ? '#dd326b' : trendDirection === 'down' ? '#17dfad' : '#888888';
                        
                        return `
                            <div style="padding: 15px; min-width: 320px;">
                                <div style="margin-bottom: 12px; font-weight: bold; color: #00d4ff; font-size: 16px;">ğŸ• ${{time}}</div>
                                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 12px;">
                                    <div>ğŸ“ˆ å¼€ç›˜: <span style="color: #fff; font-weight: bold;">$${{kline[0].toFixed(2)}}</span></div>
                                    <div>ğŸ“‰ æ”¶ç›˜: <span style="color: ${{changeColor}}; font-weight: bold;">$${{kline[1].toFixed(2)}}</span></div>
                                    <div>â¬†ï¸ æœ€é«˜: <span style="color: #00da3c; font-weight: bold;">$${{kline[3].toFixed(2)}}</span></div>
                                    <div>â¬‡ï¸ æœ€ä½: <span style="color: #ec0000; font-weight: bold;">$${{kline[2].toFixed(2)}}</span></div>
                                </div>
                                <div style="margin-bottom: 12px; padding: 8px; background: rgba(0, 212, 255, 0.15); border-radius: 6px; text-align: center;">
                                    ğŸ“Š æ¶¨è·Œ: <span style="color: ${{changeColor}}; font-weight: bold; font-size: 15px;">$${{change.toFixed(2)}} (${{changePercent}}%)</span>
                                </div>
                                <hr style="margin: 12px 0; border-color: #444;">
                                <div style="display: grid; grid-template-columns: 1fr; gap: 8px;">
                                    <div style="color: ${{trendColor}}; font-size: 14px;">ğŸ¯ VIDYAè¶‹åŠ¿: <span style="font-weight: bold;">${{trendDirection === 'up' ? 'åšå¤š â–²' : trendDirection === 'down' ? 'åšç©º â–¼' : 'ä¸­æ€§'}}</span></div>
                                    ${{upperBand !== null ? `<div style="color: #ffaa00; font-size: 13px;">ğŸ“ VIDYAä¸Šè½¨: <span style="font-weight: bold;">$${{upperBand.toFixed(2)}}</span></div>` : ''}}
                                    ${{lowerBand !== null ? `<div style="color: #00aaff; font-size: 13px;">ğŸ“ VIDYAä¸‹è½¨: <span style="font-weight: bold;">$${{lowerBand.toFixed(2)}}</span></div>` : ''}}
                                    ${{cmo !== null ? `<div style="color: #00d4ff; font-size: 13px;">ğŸ’« CMOåŠ¨é‡: <span style="font-weight: bold;">${{cmo.toFixed(2)}}</span></div>` : ''}}
                                    <div style="color: #00da3c; font-size: 13px;">ğŸ“ˆ Buy: <span style="font-weight: bold;">${{buyVolume.toLocaleString()}}</span></div>
                                    <div style="color: #ec0000; font-size: 13px;">ğŸ“‰ Sell: <span style="font-weight: bold;">${{sellVolume.toLocaleString()}}</span></div>
                                    <div style="color: ${{deltaVolumePercent >= 0 ? '#00da3c' : '#ec0000'}}; font-size: 13px;">ğŸšï¸ Delta Volume(åŠ¨æ€): <span style="font-weight: bold;">${{deltaVolumePercent >= 0 ? '+' : ''}}${{deltaVolumePercent}}%</span></div>
                                    ${{deltaVolumePercentFixed !== null && deltaVolumePercentFixed !== undefined ? `<div style="color: ${{deltaVolumePercentFixed >= 0 ? '#00da3c' : '#ec0000'}}; font-size: 14px; font-weight: bold;">âš¡ DVå›ºå®šå‘¨æœŸ(14): <span>${{deltaVolumePercentFixed >= 0 ? '+' : ''}}${{deltaVolumePercentFixed}}%</span></div>` : ''}}
                                    <hr style="margin: 8px 0; border-color: #444;">
                                    ${{ema50 !== null ? `<div style="color: #00ff00; font-size: 13px;">ğŸ“Š EMA50: <span style="font-weight: bold;">$${{ema50.toFixed(2)}}</span></div>` : ''}}
                                    ${{ema120Raw !== null ? `<div style="color: #ffff00; font-size: 13px;">ğŸ“Š EMA120(åŸ): <span style="font-weight: bold;">$${{ema120Raw.toFixed(2)}}</span></div>` : ''}}
                                    ${{ema120Smoothed !== null ? `<div style="color: #ff00ff; font-size: 13px;">ğŸ“Š EMA120(å¹³æ»‘): <span style="font-weight: bold;">$${{ema120Smoothed.toFixed(2)}}</span></div>` : ''}}
                                </div>
                            </div>
                        `;
                    }}
                }},
                grid: {{
                    left: '6%',
                    right: '6%',
                    bottom: 120,
                    top: 60,
                    backgroundColor: 'transparent'
                }},
                xAxis: {{
                    type: 'category',
                    data: data.timeData,
                    scale: true,
                    boundaryGap: false,
                    axisLine: {{ 
                        onZero: false,
                        lineStyle: {{ color: '#555' }}
                    }},
                    splitLine: {{ 
                        show: true,
                        lineStyle: {{ color: '#333', opacity: 0.6 }}
                    }},
                    axisLabel: {{
                        formatter: function (value) {{
                            return value.slice(5, 16);
                        }},
                        color: '#cccccc',
                        fontSize: 12
                    }}
                }},
                yAxis: {{
                    scale: true,
                    splitArea: {{
                        show: true,
                        areaStyle: {{
                            color: [['rgba(0, 212, 255, 0.03)', 'rgba(0, 212, 255, 0.01)']]
                        }}
                    }},
                    splitLine: {{
                        lineStyle: {{ color: '#333', opacity: 0.6 }}
                    }},
                    axisLabel: {{
                        formatter: function (value) {{
                            return '$' + value.toFixed(0);
                        }},
                        color: '#cccccc',
                        fontSize: 12
                    }},
                    axisLine: {{
                        lineStyle: {{ color: '#555' }}
                    }}
                }},
                dataZoom: [
                    {{
                        type: 'inside',
                        start: 80,
                        end: 100
                    }},
                    {{
                        show: true,
                        type: 'slider',
                        top: '94%',
                        start: 80,
                        end: 100,
                        backgroundColor: 'rgba(45, 45, 45, 0.8)',
                        borderColor: '#00d4ff',
                        fillerColor: 'rgba(0, 212, 255, 0.25)',
                        handleStyle: {{
                            color: '#00d4ff',
                            borderColor: '#00d4ff'
                        }},
                        textStyle: {{
                            color: '#cccccc'
                        }}
                    }}
                ],
                series: [
                    {{
                        name: 'Kçº¿',
                        type: 'candlestick',
                        data: data.klineData,
                        itemStyle: {{
                            color: '#00da3c',
                            color0: '#ec0000',
                            borderColor: '#00da3c',
                            borderColor0: '#ec0000',
                            borderWidth: 1.5
                        }}
                    }},
                    {{
                        name: 'VIDYAä¸Šè½¨',
                        type: 'line',
                        data: data.upperBandData || [],
                        smooth: true,
                        lineStyle: {{
                            color: '#ffaa00',  // ğŸ”´ é»„è‰²ç»†çº¿
                            width: 1,
                            type: 'solid'
                        }},
                        symbol: 'none',
                        z: 2
                    }},
                    {{
                        name: 'VIDYAä¸‹è½¨',
                        type: 'line',
                        data: data.lowerBandData || [],
                        smooth: true,
                        lineStyle: {{
                            color: '#00aaff',  // ğŸ”´ è“è‰²ç»†çº¿
                            width: 1,
                            type: 'solid'
                        }},
                        symbol: 'none',
                        z: 2
                    }},
                    {{
                        name: 'åšå¤šè¶‹åŠ¿çº¿',
                        type: 'line',
                        data: (function() {{
                            // ğŸ”´ åšå¤šæ—¶æ˜¾ç¤ºä¸‹è½¨ç²—çº¿ï¼Œåšç©ºæ—¶éšè—
                            const lowerBand = data.lowerBandData || [];
                            const trendDirection = data.trendDirection || [];
                            const result = [];
                            
                            for (let i = 0; i < trendDirection.length; i++) {{
                                if (trendDirection[i] === 'up') {{
                                    // åšå¤šæ—¶æ˜¾ç¤ºä¸‹è½¨
                                    result.push(lowerBand[i]);
                                }} else {{
                                    // åšç©ºæ—¶éšè—
                                    result.push(null);
                                }}
                            }}
                            return result;
                        }})(),
                        smooth: true,
                        lineStyle: {{
                            color: '#ff6600',  // ğŸ”´ æ©™è‰²ç²—çº¿
                            width: 3,
                            type: 'solid'
                        }},
                        symbol: 'none',
                        connectNulls: false,
                        z: 3
                    }},
                    {{
                        name: 'åšç©ºè¶‹åŠ¿çº¿',
                        type: 'line',
                        data: (function() {{
                            // ğŸ”´ åšç©ºæ—¶æ˜¾ç¤ºä¸Šè½¨ç²—çº¿ï¼Œåšå¤šæ—¶éšè—
                            const upperBand = data.upperBandData || [];
                            const trendDirection = data.trendDirection || [];
                            const result = [];
                            
                            for (let i = 0; i < trendDirection.length; i++) {{
                                if (trendDirection[i] === 'down') {{
                                    // åšç©ºæ—¶æ˜¾ç¤ºä¸Šè½¨
                                    result.push(upperBand[i]);
                                }} else {{
                                    // åšå¤šæ—¶éšè—
                                    result.push(null);
                                }}
                            }}
                            return result;
                        }})(),
                        smooth: true,
                        lineStyle: {{
                            color: '#0066ff',  // ğŸ”´ è“è‰²ç²—çº¿
                            width: 3,
                            type: 'solid'
                        }},
                        symbol: 'none',
                        connectNulls: false,
                        z: 3
                    }},
                    {{
                        name: 'æ”¯æ’‘ä½',
                        type: 'custom',
                        renderItem: function(params, api) {{
                            const supportData = data.supportLevels || [];
                            const dataIndex = params.dataIndex;
                            const supportList = supportData[dataIndex];  // ç°åœ¨æ˜¯æ•°ç»„
                            
                            // å¦‚æœæ²¡æœ‰æ”¯æ’‘ä½æ•°æ®ï¼Œè¿”å›null
                            if (!supportList || !Array.isArray(supportList) || supportList.length === 0) {{
                                return null;
                            }}
                            
                            // ğŸ”´ ä¸ºæ¯ä¸ªæ”¯æ’‘ä½ç”»ä¸€æ¡çŸ­çº¿ï¼ˆ2æ ¹Kçº¿å®½åº¦ï¼‰
                            const lineLength = 2;
                            const lines = [];
                            
                            for (let i = 0; i < supportList.length; i++) {{
                                const supportValue = supportList[i];
                                
                                // æ£€æŸ¥æ˜¯å¦æ˜¯æ–°çš„æ”¯æ’‘ä½ï¼ˆä¸å‰ä¸€ä¸ªæ—¶é—´ç‚¹çš„ä¸åŒï¼‰
                                const prevSupportList = dataIndex > 0 ? supportData[dataIndex - 1] : [];
                                const isNewSupport = !prevSupportList || !prevSupportList.includes(supportValue);
                                
                                if (!isNewSupport) {{
                                    continue;  // ä¸æ˜¯æ–°æ”¯æ’‘ä½ï¼Œè·³è¿‡
                                }}
                                
                                const startIndex = dataIndex;
                                const endIndex = Math.min(dataIndex + lineLength, supportData.length - 1);
                                
                                const startPoint = api.coord([startIndex, supportValue]);
                                const endPoint = api.coord([endIndex, supportValue]);
                                
                                lines.push({{
                                    type: 'line',
                                    shape: {{
                                        x1: startPoint[0],
                                        y1: startPoint[1],
                                        x2: endPoint[0],
                                        y2: endPoint[1]
                                    }},
                                    style: {{
                                        stroke: '#17dfad',  // ç»¿è‰²
                                        lineWidth: 1,
                                        opacity: 0.8
                                    }}
                                }});
                            }}
                            
                            // å¦‚æœæœ‰å¤šæ¡çº¿ï¼Œè¿”å›groupï¼›å¦åˆ™è¿”å›å•æ¡çº¿
                            if (lines.length === 0) {{
                                return null;
                            }} else if (lines.length === 1) {{
                                return lines[0];
                            }} else {{
                                return {{
                                    type: 'group',
                                    children: lines
                                }};
                            }}
                        }},
                        data: data.supportLevels.map((value, index) => [index, value]),
                        z: 5
                    }},
                    {{
                        name: 'é˜»åŠ›ä½',
                        type: 'custom',
                        renderItem: function(params, api) {{
                            const resistanceData = data.resistanceLevels || [];
                            const dataIndex = params.dataIndex;
                            const resistanceList = resistanceData[dataIndex];  // ç°åœ¨æ˜¯æ•°ç»„
                            
                            // å¦‚æœæ²¡æœ‰é˜»åŠ›ä½æ•°æ®ï¼Œè¿”å›null
                            if (!resistanceList || !Array.isArray(resistanceList) || resistanceList.length === 0) {{
                                return null;
                            }}
                            
                            // ğŸ”´ ä¸ºæ¯ä¸ªé˜»åŠ›ä½ç”»ä¸€æ¡çŸ­çº¿ï¼ˆ2æ ¹Kçº¿å®½åº¦ï¼‰
                            const lineLength = 2;
                            const lines = [];
                            
                            for (let i = 0; i < resistanceList.length; i++) {{
                                const resistanceValue = resistanceList[i];
                                
                                // æ£€æŸ¥æ˜¯å¦æ˜¯æ–°çš„é˜»åŠ›ä½ï¼ˆä¸å‰ä¸€ä¸ªæ—¶é—´ç‚¹çš„ä¸åŒï¼‰
                                const prevResistanceList = dataIndex > 0 ? resistanceData[dataIndex - 1] : [];
                                const isNewResistance = !prevResistanceList || !prevResistanceList.includes(resistanceValue);
                                
                                if (!isNewResistance) {{
                                    continue;  // ä¸æ˜¯æ–°é˜»åŠ›ä½ï¼Œè·³è¿‡
                                }}
                                
                                const startIndex = dataIndex;
                                const endIndex = Math.min(dataIndex + lineLength, resistanceData.length - 1);
                                
                                const startPoint = api.coord([startIndex, resistanceValue]);
                                const endPoint = api.coord([endIndex, resistanceValue]);
                                
                                lines.push({{
                                    type: 'line',
                                    shape: {{
                                        x1: startPoint[0],
                                        y1: startPoint[1],
                                        x2: endPoint[0],
                                        y2: endPoint[1]
                                    }},
                                    style: {{
                                        stroke: '#dd326b',  // çº¢è‰²
                                        lineWidth: 1,
                                        opacity: 0.8
                                    }}
                                }});
                            }}
                            
                            // å¦‚æœæœ‰å¤šæ¡çº¿ï¼Œè¿”å›groupï¼›å¦åˆ™è¿”å›å•æ¡çº¿
                            if (lines.length === 0) {{
                                return null;
                            }} else if (lines.length === 1) {{
                                return lines[0];
                            }} else {{
                                return {{
                                    type: 'group',
                                    children: lines
                                }};
                            }}
                        }},
                        data: data.resistanceLevels.map((value, index) => [index, value]),
                        z: 5
                    }},
                    {{
                        name: 'EMA50',
                        type: 'line',
                        data: data.ema50Data || [],
                        smooth: true,
                        lineStyle: {{
                            color: '#00ff00',  // ğŸ”´ ç»¿è‰²
                            width: 2,
                            type: 'solid'
                        }},
                        symbol: 'none',
                        z: 4
                    }},
                    {{
                        name: 'EMA120åŸå§‹',
                        type: 'line',
                        data: data.ema120RawData || [],
                        smooth: true,
                        lineStyle: {{
                            color: '#ffff00',  // ğŸ”´ é»„è‰²
                            width: 1,
                            type: 'dashed'  // è™šçº¿
                        }},
                        symbol: 'none',
                        z: 3
                    }},
                    {{
                        name: 'EMA120å¹³æ»‘',
                        type: 'line',
                        data: data.ema120SmoothedData || [],
                        smooth: true,
                        lineStyle: {{
                            color: '#ff00ff',  // ğŸ”´ ç´«è‰²
                            width: 2,
                            type: 'solid'
                        }},
                        symbol: 'none',
                        z: 4
                    }}
                ]
            }};
        }}
        
        function updateInfoPanel(time, kline, vidya, smoothedVidya, upperBand, lowerBand, support, resistance) {{
            document.getElementById('currentTime').textContent = time;
            document.getElementById('currentOpen').textContent = `$${{kline[0].toFixed(2)}}`;
            document.getElementById('currentHigh').textContent = `$${{kline[3].toFixed(2)}}`;
            document.getElementById('currentLow').textContent = `$${{kline[2].toFixed(2)}}`;
            document.getElementById('currentClose').textContent = `$${{kline[1].toFixed(2)}}`;
            
            // ğŸ”´ è·å–å½“å‰è¶‹åŠ¿å’Œäº¤æ˜“é‡ä¿¡æ¯
            const dataIndex = backtestData.timeData.indexOf(time);
            if (dataIndex >= 0) {{
                const trendDirection = backtestData.trendDirection[dataIndex];
                const volume = backtestData.volumeData[dataIndex];
                const deltaVolume = backtestData.deltaVolumeData[dataIndex];
                const upperBand = backtestData.upperBandData[dataIndex];
                const lowerBand = backtestData.lowerBandData[dataIndex];
                
                // ğŸ”´ ä»åç«¯è·å–çœŸå®çš„ç´¯ç§¯Buy/Selläº¤æ˜“é‡
                const buyVolume = backtestData.buyVolumeData[dataIndex] || 0;
                const sellVolume = backtestData.sellVolumeData[dataIndex] || 0;
                const avgVolume = (buyVolume + sellVolume) / 2;
                const deltaVolumePercent = avgVolume > 0 ? ((buyVolume - sellVolume) / avgVolume * 100).toFixed(2) : '0.00';
                
                document.getElementById('currentTrend').textContent = trendDirection === 'up' ? 'åšå¤š â–²' : trendDirection === 'down' ? 'åšç©º â–¼' : 'ä¸­æ€§';
                document.getElementById('currentUpperBand').textContent = upperBand !== null ? `$${{upperBand.toFixed(2)}}` : 'N/A';
                document.getElementById('currentLowerBand').textContent = lowerBand !== null ? `$${{lowerBand.toFixed(2)}}` : 'N/A';
                document.getElementById('currentBuyVolume').textContent = buyVolume.toLocaleString();
                document.getElementById('currentSellVolume').textContent = sellVolume.toLocaleString();
                document.getElementById('currentDeltaVolume').textContent = `${{deltaVolumePercent >= 0 ? '+' : ''}}${{deltaVolumePercent}}%`;
                
                // ğŸ”´ è·å–CMOæ•°æ®
                const cmo = backtestData.cmoData[dataIndex];
                document.getElementById('currentCMO').textContent = cmo !== null ? `${{cmo.toFixed(2)}}` : 'N/A';
                
                // ğŸ”´ è·å–å›ºå®šå‘¨æœŸDelta Volumeç™¾åˆ†æ¯”
                const deltaVolumePercentFixed = backtestData.deltaVolumePercentFixed[dataIndex];
                const deltaVolumeFixedText = deltaVolumePercentFixed !== null && deltaVolumePercentFixed !== undefined 
                    ? `${{deltaVolumePercentFixed >= 0 ? '+' : ''}}${{deltaVolumePercentFixed}}%` 
                    : 'N/A';
                const deltaVolumeFixedColor = deltaVolumePercentFixed >= 0 ? '#00da3c' : '#ec0000';
                const deltaVolumeFixedElement = document.getElementById('currentDeltaVolumeFixed');
                deltaVolumeFixedElement.textContent = deltaVolumeFixedText;
                deltaVolumeFixedElement.style.color = deltaVolumePercentFixed !== null && deltaVolumePercentFixed !== undefined ? deltaVolumeFixedColor : '#fff';
                
                // ğŸ”´ è·å–EMAæ•°æ®
                const ema50 = backtestData.ema50Data[dataIndex];
                const ema120Raw = backtestData.ema120RawData[dataIndex];
                const ema120Smoothed = backtestData.ema120SmoothedData[dataIndex];
                
                document.getElementById('currentEma50').textContent = ema50 !== null ? `$${{ema50.toFixed(2)}}` : 'N/A';
                document.getElementById('currentEma120Raw').textContent = ema120Raw !== null ? `$${{ema120Raw.toFixed(2)}}` : 'N/A';
                document.getElementById('currentEma120Smoothed').textContent = ema120Smoothed !== null ? `$${{ema120Smoothed.toFixed(2)}}` : 'N/A';
            }} else {{
                document.getElementById('currentTrend').textContent = 'N/A';
                document.getElementById('currentUpperBand').textContent = 'N/A';
                document.getElementById('currentLowerBand').textContent = 'N/A';
                document.getElementById('currentCMO').textContent = 'N/A';
                document.getElementById('currentBuyVolume').textContent = 'N/A';
                document.getElementById('currentSellVolume').textContent = 'N/A';
                document.getElementById('currentDeltaVolume').textContent = 'N/A';
                document.getElementById('currentDeltaVolumeFixed').textContent = 'N/A';
                document.getElementById('currentEma50').textContent = 'N/A';
                document.getElementById('currentEma120Raw').textContent = 'N/A';
                document.getElementById('currentEma120Smoothed').textContent = 'N/A';
            }}
        }}
        
        myChart.setOption(getChartOption(backtestData));
        
        myChart.on('click', function (params) {{
            if (params.componentType === 'series') {{
                const dataIndex = params.dataIndex;
                const time = backtestData.timeData[dataIndex];
                const kline = backtestData.klineData[dataIndex];
                
                updateInfoPanel(time, kline, null, null, null, null, null, null, null);
            }}
        }});
        
        window.addEventListener('resize', function() {{
            myChart.resize();
        }});
        
        console.log('ğŸš€ çœŸå®å›æµ‹æ•°æ®å›¾è¡¨å·²åŠ è½½');
    </script>
</body>
</html>'''
    
    # ä¿å­˜HTMLæ–‡ä»¶åˆ°æŒ‡å®šç›®å½•
    html_file = os.path.join(output_dir, "äº¤äº’å¼å›¾è¡¨.html")
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"ğŸŒ äº¤äº’å¼å›¾è¡¨å·²ä¿å­˜: {html_file}")
    return html_file

def main():
    print("å¼€å§‹æµ‹è¯•çº¯VIDYAç­–ç•¥...")
    
    # ä½¿ç”¨é…ç½®æ–‡ä»¶ä¸­çš„æ•°æ®åº“å‚æ•°åˆå§‹åŒ–db_service
    db_service = DatabaseService(**LOCAL_DATABASE_CONFIG)

    # è·å–ç­–ç•¥é…ç½®
    config = get_strategy_config()

    # æ˜¾ç¤ºé…ç½®ä¿¡æ¯
    print_config_info()
    
    # åˆ›å»ºçº¯VIDYAç­–ç•¥å®ä¾‹
    strategy = TrendVolumaticDynamicAverageStrategy(
        timeframe=config['timeframe'],
        initial_capital=config['initial_capital'],
        position_size_percentage=config['position_size_percentage'],
        fixed_take_profit_pct=config['fixed_take_profit_pct'],
        max_loss_pct=config['max_loss_pct'],
        volatility_timeframe=config['volatility_timeframe'],
        volatility_length=config['volatility_length'],
        volatility_mult=config['volatility_mult'],
        volatility_ema_period=config['volatility_ema_period'],
        volatility_threshold=config['volatility_threshold'],
        basis_change_threshold=config['basis_change_threshold'],
        # ğŸ”´ æ ‡å‡†VIDYAå‚æ•°
        vidya_length=config.get('vidya_length', 20),
        vidya_momentum=config.get('vidya_momentum', 9),
        vidya_smooth=config.get('vidya_smooth', 15),
        vidya_band_distance=config.get('vidya_band_distance', 2.0),
        vidya_atr_period=config.get('vidya_atr_period', 200),
        vidya_pivot_left=config.get('vidya_pivot_left', 3),
        vidya_pivot_right=config.get('vidya_pivot_right', 3),
        delta_volume_period=config.get('delta_volume_period', 14),  # ğŸ”´ å›ºå®šå‘¨æœŸDelta Volume
        # ğŸ”´ å¼€ä»“æ¡ä»¶é…ç½®ï¼ˆç‹¬ç«‹å¼€å…³ï¼‰
        entry_condition_trend_breakthrough=config.get('entry_condition_trend_breakthrough', True),
        entry_condition_arrow_signal=config.get('entry_condition_arrow_signal', False),
        entry_condition_vidya_slope=config.get('entry_condition_vidya_slope', False),
        entry_condition_delta_volume=config.get('entry_condition_delta_volume', True),
        entry_condition_ema_120_slope=config.get('entry_condition_ema_120_slope', False),
        # ğŸ“ å¸ƒæ—å¸¦ä¸­è½¨è§’åº¦è®¡ç®—å™¨å‚æ•°
        enable_bb_angle_entry=config.get('enable_bb_angle_entry', False),
        bb_midline_period=config.get('bb_midline_period', 14),
        bb_angle_window_size=config.get('bb_angle_window_size', 14),
        bb_angle_threshold=config.get('bb_angle_threshold', 0.3),
        bb_r_squared_threshold=config.get('bb_r_squared_threshold', 0.6),
        bb_stop_loss_lock_periods=config.get('bb_stop_loss_lock_periods', 5),
        bb_max_loss_pct=config.get('bb_max_loss_pct', 1.0)  # ğŸ”´ å¸ƒæ—å¸¦è§’åº¦å¼€ä»“çš„æœ€å¤§äºæŸç™¾åˆ†æ¯”
    )

    print(f"\nçº¯VIDYAç­–ç•¥åˆå§‹åŒ–å®Œæˆ")
    print(f"ğŸ“Š æ—¶é—´å‘¨æœŸ: {config['timeframe']}")
    print(f"ğŸ¯ VIDYAå‚æ•°: length={config.get('vidya_length', 20)}, momentum={config.get('vidya_momentum', 9)}, smooth={config.get('vidya_smooth', 15)}")
    print(f"ğŸ“ ATRå¸¦å®½: distance={config.get('vidya_band_distance', 2.0)}, period={config.get('vidya_atr_period', 200)}")
    print(f"ğŸ¯ äº¤æ˜“é€»è¾‘: VIDYAè¶‹åŠ¿æ”¹å˜å¼€ä»“ â†’ å›ºå®šæ­¢ç›ˆ{config['fixed_take_profit_pct']}% + VIDYAåŠ¨æ€æ­¢æŸ")
    
    # äº¤æ˜“ç»Ÿè®¡
    trades = []
    initial_capital = config['initial_capital']
    current_position = None

    df = db_service.get_kline_data(config['long_coin'], config['start_date'], config['end_date'])

    if df.empty:
        print("æœªè·å–åˆ°æ•°æ®ï¼Œè¯·æ£€æŸ¥æ•°æ®åº“è¿æ¥å’Œè¡¨ç»“æ„")
        return

    print(f"\nğŸ“Š è·å–åˆ° {len(df)} ä¸ªæ•°æ®ç‚¹")
    print(f"æ—¶é—´èŒƒå›´: {df['timestamp'].min()} åˆ° {df['timestamp'].max()}")
    print(f"ä»·æ ¼èŒƒå›´: ${df['low'].min():.2f} - ${df['high'].max():.2f}")

    # ğŸ”¥ æ»¤æ³¢å™¨é¢„çƒ­ï¼šè·å–å›æµ‹å¼€å§‹å‰çš„å†å²æ•°æ®
    start_timestamp = pd.to_datetime(config['start_date'])
    warmup_days = 25  # ğŸ”¥ æ•°æ®é¢„çƒ­ï¼ˆç¡®ä¿VIDYAæŒ‡æ ‡å®Œå…¨ç¨³å®šï¼Œæ¥è¿‘TradingViewæ•ˆæœï¼‰
    warmup_start = start_timestamp - pd.Timedelta(days=warmup_days)
    warmup_start_str = warmup_start.strftime('%Y-%m-%d %H:%M:%S')
    warmup_end_str = config['start_date']
    
    print(f"\nğŸ”¥ è·å–é¢„çƒ­æ•°æ®...")
    print(f"é¢„çƒ­æ—¶é—´èŒƒå›´: {warmup_start_str} åˆ° {warmup_end_str}")
    
    warmup_df = db_service.get_kline_data(config['long_coin'], warmup_start_str, warmup_end_str)
    
    if not warmup_df.empty:
        print(f"ğŸ“ˆ é¢„çƒ­æ•°æ®: {len(warmup_df)} ä¸ªæ•°æ®ç‚¹")
        
        # å‡†å¤‡é¢„çƒ­æ•°æ®ï¼ˆåŒ…å«å®Œæ•´çš„OHLCã€volumeå’Œæ—¶é—´æˆ³ï¼‰
        warmup_data = []
        for _, row in warmup_df.iterrows():
            warmup_data.append({
                'timestamp': row['timestamp'],
                'open': row['open'],
                'high': row['high'],
                'low': row['low'],
                'close': row['close'],
                'volume': row.get('volume', 0)  # ğŸ”´ æ·»åŠ æˆäº¤é‡
            })
        
        # æ‰§è¡Œé¢„çƒ­
        strategy.warmup_filter(warmup_data)
    else:
        print("âš ï¸  æœªè·å–åˆ°é¢„çƒ­æ•°æ®ï¼Œå°†ä½¿ç”¨é»˜è®¤åˆå§‹åŒ–")

    print(f"\nğŸš€ å¼€å§‹æ­£å¼å›æµ‹ï¼Œå…± {len(df)} ä¸ªæ•°æ®ç‚¹")

    # ğŸ¨ åˆå§‹åŒ–å›¾è¡¨æ•°æ®æ”¶é›†ï¼ˆçº¯VIDYAï¼‰
    chart_data = {
        'timeData': [],
        'klineData': [],
        'vidyaData': [],  # ğŸ”´ VIDYAæ•°æ®
        'vidyaSmoothedData': [],  # ğŸ”´ å¹³æ»‘VIDYAæ•°æ®
        'cmoData': [],  # ğŸ”´ CMOæ•°æ®
        'volumeData': [],  # ğŸ”´ æˆäº¤é‡æ•°æ®
        'buyVolumeData': [],  # ğŸ”´ Buy Volumeæ•°æ®ï¼ˆç´¯ç§¯ï¼‰
        'sellVolumeData': [],  # ğŸ”´ Sell Volumeæ•°æ®ï¼ˆç´¯ç§¯ï¼‰
        'deltaVolumeData': [],  # ğŸ”´ Delta Volumeæ•°æ®ï¼ˆåŠ¨æ€ç´¯ç§¯ï¼‰
        'deltaVolumePercentFixed': [],  # ğŸ”´ å›ºå®šå‘¨æœŸDelta Volumeç™¾åˆ†æ¯”ï¼ˆæ¯1åˆ†é’Ÿæ›´æ–°ï¼‰
        'upperBandData': [],  # ğŸ”´ ATRä¸Šè½¨æ•°æ®
        'lowerBandData': [],  # ğŸ”´ ATRä¸‹è½¨æ•°æ®
        'atrData': [],  # ğŸ”´ ATRæ•°æ®
        'supportLevels': [],  # ğŸ”´ æ”¯æ’‘çº¿æ•°æ®
        'resistanceLevels': [],  # ğŸ”´ é˜»åŠ›çº¿æ•°æ®
        'trendDirection': [],  # ğŸ”´ è¶‹åŠ¿æ–¹å‘æ•°æ® ('up', 'down', 'neutral')
        'trendChangeSignals': [],  # ğŸ”´ è¶‹åŠ¿æ”¹å˜ä¿¡å·ï¼ˆç”¨äºæ˜¾ç¤ºç®­å¤´ï¼‰
        'ema50Data': [],  # ğŸ”´ EMA50æ•°æ®
        'ema120RawData': [],  # ğŸ”´ EMA120åŸå§‹æ•°æ®
        'ema120SmoothedData': [],  # ğŸ”´ EMA120å¹³æ»‘æ•°æ®ï¼ˆSMA50ï¼‰
        'tradeSignals': [],
        'config_info': {
            'start_date': config['start_date'],
            'end_date': config['end_date'],
            'timeframe': config['timeframe'],
            'initial_capital': config['initial_capital'],
            'vidya_length': config.get('vidya_length', 20),
            'vidya_momentum': config.get('vidya_momentum', 9),
            'vidya_smooth': config.get('vidya_smooth', 15),
            'vidya_band_distance': config.get('vidya_band_distance', 2.0),
            'vidya_atr_period': config.get('vidya_atr_period', 200),
            'vidya_pivot_left': config.get('vidya_pivot_left', 3),
            'vidya_pivot_right': config.get('vidya_pivot_right', 3),
            'use_vidya_trading': config.get('use_vidya_trading', True),
            'use_delta_volume_filter': config.get('use_delta_volume_filter', True)
        }
    }
    
    # Kçº¿èšåˆçŠ¶æ€ï¼ˆåŠ¨æ€æ ¹æ®é…ç½®è°ƒæ•´ï¼‰
    current_kline = None
    current_kline_start = None

    # æ‰§è¡Œç­–ç•¥
    for index, row in df.iterrows():
        timestamp = row['timestamp']
        open_price = row['open']
        high_price = row['high']
        low_price = row['low']
        close_price = row['close']
        volume = row.get('volume', 0)  # ğŸ”´ è·å–æˆäº¤é‡

        # ğŸ¨ æ”¶é›†Kçº¿æ•°æ®ï¼ˆç”¨äºå›¾è¡¨ï¼Œæ ¹æ®é…ç½®åŠ¨æ€è°ƒæ•´ï¼‰
        timeframe = config['timeframe']
        
        # æ ¹æ®æ—¶é—´å‘¨æœŸè®¡ç®—Kçº¿å¼€å§‹æ—¶é—´
        if timeframe == '15m':
            minute = timestamp.minute
            if minute < 15:
                period_minute = 0
            elif minute < 30:
                period_minute = 15
            elif minute < 45:
                period_minute = 30
            else:
                period_minute = 45
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        elif timeframe == '20m':
            minute = timestamp.minute
            if minute < 20:
                period_minute = 0
            elif minute < 40:
                period_minute = 20
            else:
                period_minute = 40
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        elif timeframe == '30m':
            minute = timestamp.minute
            if minute < 30:
                period_minute = 0
            else:
                period_minute = 30
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        elif timeframe == '1h':
            kline_start = timestamp.replace(minute=0, second=0, microsecond=0)
        elif timeframe == '2h':
            hour = timestamp.hour
            period_hour = (hour // 2) * 2
            kline_start = timestamp.replace(hour=period_hour, minute=0, second=0, microsecond=0)
        elif timeframe == '4h':
            hour = timestamp.hour
            period_hour = (hour // 4) * 4
            kline_start = timestamp.replace(hour=period_hour, minute=0, second=0, microsecond=0)
        else:
            # é»˜è®¤30åˆ†é’Ÿ
            minute = timestamp.minute
            if minute < 30:
                period_minute = 0
            else:
                period_minute = 30
            kline_start = timestamp.replace(minute=period_minute, second=0, microsecond=0)
        
        # æ£€æŸ¥æ˜¯å¦æ˜¯æ–°çš„Kçº¿å‘¨æœŸ
        if current_kline_start is None or kline_start != current_kline_start:
            # ä¿å­˜ä¸Šä¸€ä¸ªKçº¿å’ŒæŒ‡æ ‡æ•°æ®
            if current_kline is not None:
                chart_data['timeData'].append(current_kline_start.strftime('%Y-%m-%d %H:%M:%S'))
                chart_data['klineData'].append([
                    round(current_kline['open'], 2),
                    round(current_kline['close'], 2),
                    round(current_kline['low'], 2),
                    round(current_kline['high'], 2)
                ])
                
                # SARç›¸å…³æ•°æ®å·²åˆ é™¤ï¼Œä¸å†æ”¶é›†
                
                # ğŸ”´ æ”¶é›†VIDYAæ•°æ®
                current_vidya = strategy.vidya_indicator.current_vidya
                if current_vidya is not None:
                    chart_data['vidyaData'].append(round(current_vidya, 2))
                else:
                    chart_data['vidyaData'].append(None)
                
                # ğŸ”´ æ”¶é›†å¹³æ»‘VIDYAæ•°æ®
                if len(strategy.vidya_indicator.smoothed_vidya_values) > 0:
                    current_smoothed_vidya = strategy.vidya_indicator.smoothed_vidya_values[-1]
                    chart_data['vidyaSmoothedData'].append(round(current_smoothed_vidya, 2))
                else:
                    chart_data['vidyaSmoothedData'].append(None)
                
                # ğŸ”´ æ”¶é›†CMOæ•°æ®
                if len(strategy.vidya_indicator.close_history) >= strategy.vidya_indicator.vidya_momentum + 1:
                    # è®¡ç®—å½“å‰CMOå€¼
                    current_cmo = strategy._calculate_cmo(
                        strategy.vidya_indicator.close_history, 
                        strategy.vidya_indicator.vidya_momentum
                    )
                    chart_data['cmoData'].append(round(current_cmo, 2))
                else:
                    chart_data['cmoData'].append(None)
                
                # ğŸ”´ æ”¶é›†æˆäº¤é‡æ•°æ®
                chart_data['volumeData'].append(current_kline.get('volume', 0))
                
                # ğŸ”´ æ”¶é›†Buy/Sell Volumeï¼ˆç´¯ç§¯å€¼ï¼‰
                current_buy_volume = strategy.vidya_indicator.buy_volume
                current_sell_volume = strategy.vidya_indicator.sell_volume
                chart_data['buyVolumeData'].append(current_buy_volume)
                chart_data['sellVolumeData'].append(current_sell_volume)
                
                # ğŸ”´ æ”¶é›†Delta Volumeï¼ˆåŠ¨æ€ç´¯ç§¯ï¼‰
                current_delta_volume = strategy.vidya_indicator.delta_volume
                chart_data['deltaVolumeData'].append(current_delta_volume)
                
                # ğŸ”´ æ”¶é›†å›ºå®šå‘¨æœŸDelta Volumeç™¾åˆ†æ¯”ï¼ˆæ¯1åˆ†é’Ÿæ›´æ–°ï¼‰
                current_delta_volume_percent_fixed = strategy.vidya_indicator.delta_volume_percent_fixed
                chart_data['deltaVolumePercentFixed'].append(round(current_delta_volume_percent_fixed, 2))
                
                # ğŸ”´ æ”¶é›†ATRå¸¦å®½æ•°æ®ï¼ˆä½¿ç”¨ç­–ç•¥ä¿å­˜çš„å½“å‰å€¼ï¼Œç¡®ä¿ä¸å¹³ä»“é€»è¾‘ä¸€è‡´ï¼‰
                current_upper_band = strategy.current_upper_band
                current_lower_band = strategy.current_lower_band
                
                if current_upper_band is not None:
                    chart_data['upperBandData'].append(round(current_upper_band, 2))
                else:
                    chart_data['upperBandData'].append(None)
                
                if current_lower_band is not None:
                    chart_data['lowerBandData'].append(round(current_lower_band, 2))
                else:
                    chart_data['lowerBandData'].append(None)
                
                # ğŸ”´ æ”¶é›†ATRæ•°æ®
                if len(strategy.vidya_indicator.atr_values) > 0:
                    current_atr = strategy.vidya_indicator.atr_values[-1]
                    chart_data['atrData'].append(round(current_atr, 2))
                else:
                    chart_data['atrData'].append(None)
                
                # ğŸ”´ æ”¶é›†æ‰€æœ‰æ”¯æ’‘é˜»åŠ›çº¿æ•°æ®ï¼ˆç”¨äºå›¾è¡¨æ˜¾ç¤ºï¼‰
                # è·å–å½“å‰æ‰€æœ‰çš„æ”¯æ’‘é˜»åŠ›ä½ï¼ˆæœ€å¤š3ä¸ªï¼‰
                all_support_levels = strategy.vidya_indicator.support_levels if strategy.vidya_indicator.support_levels else []
                all_resistance_levels = strategy.vidya_indicator.resistance_levels if strategy.vidya_indicator.resistance_levels else []
                
                # å°†æ‰€æœ‰æ”¯æ’‘é˜»åŠ›ä½å­˜å‚¨ä¸ºåˆ—è¡¨ï¼ˆå›¾è¡¨ä¼šæ˜¾ç¤ºæ‰€æœ‰çš„çº¿ï¼‰
                support_list = [round(s, 2) for s in all_support_levels] if all_support_levels else []
                resistance_list = [round(r, 2) for r in all_resistance_levels] if all_resistance_levels else []
                
                chart_data['supportLevels'].append(support_list)
                chart_data['resistanceLevels'].append(resistance_list)
                
                # ğŸ”´ æ”¶é›†EMAæ•°æ®
                if strategy.vidya_indicator.current_ema_50 is not None:
                    chart_data['ema50Data'].append(round(strategy.vidya_indicator.current_ema_50, 2))
                else:
                    chart_data['ema50Data'].append(None)
                
                if strategy.vidya_indicator.current_ema_120 is not None:
                    chart_data['ema120RawData'].append(round(strategy.vidya_indicator.current_ema_120, 2))
                else:
                    chart_data['ema120RawData'].append(None)
                
                if strategy.vidya_indicator.current_ema_120_smoothed is not None:
                    chart_data['ema120SmoothedData'].append(round(strategy.vidya_indicator.current_ema_120_smoothed, 2))
                else:
                    chart_data['ema120SmoothedData'].append(None)
                
                # ğŸ”´ æ”¶é›†è¶‹åŠ¿æ–¹å‘æ•°æ®
                current_trend = strategy.vidya_indicator.current_trend
                previous_trend = strategy.vidya_indicator.previous_trend
                chart_data['trendDirection'].append(current_trend if current_trend else 'neutral')
                
                # ğŸ”´ æ£€æµ‹è¶‹åŠ¿æ”¹å˜å¹¶è®°å½•ç®­å¤´ä¿¡å·
                if previous_trend is not None and current_trend != previous_trend and current_trend != 'neutral':
                    # è¶‹åŠ¿æ”¹å˜äº†ï¼Œæ·»åŠ ç®­å¤´æ ‡è®°
                    signal_data = {
                        'time': current_kline_start.strftime('%Y-%m-%d %H:%M:%S'),
                        'price': round(current_kline['close'], 2),
                        'direction': current_trend,  # 'up' or 'down'
                        'from_trend': previous_trend,
                        'to_trend': current_trend
                    }
                    chart_data['trendChangeSignals'].append(signal_data)
                    print(f"    ğŸ”„ æ£€æµ‹åˆ°VIDYAè¶‹åŠ¿æ”¹å˜: {previous_trend} â†’ {current_trend} @ {current_kline_start.strftime('%H:%M')} ä»·æ ¼${current_kline['close']:.2f}")
            
            # å¼€å§‹æ–°çš„Kçº¿å‘¨æœŸ
            current_kline_start = kline_start
            current_kline = {
                'open': open_price,
                'high': high_price,
                'low': low_price,
                'close': close_price,
                'volume': volume  # ğŸ”´ åˆå§‹åŒ–æˆäº¤é‡
            }
        else:
            # æ›´æ–°å½“å‰Kçº¿
            current_kline['high'] = max(current_kline['high'], high_price)
            current_kline['low'] = min(current_kline['low'], low_price)
            current_kline['close'] = close_price
            current_kline['volume'] = current_kline.get('volume', 0) + volume  # ğŸ”´ ç´¯åŠ æˆäº¤é‡

        # æ›´æ–°è¶‹åŠ¿æ»¤æ³¢å™¨ç­–ç•¥ï¼ˆä¼ é€’volumeï¼‰
        result = strategy.update(timestamp, open_price, high_price, low_price, close_price, volume)
        
        # å¤„ç†äº¤æ˜“ä¿¡å·
        for signal in result['signals']:
            # ğŸ”§ ä½¿ç”¨æ›´ç²¾ç¡®çš„æ—¶é—´æˆ³ï¼šå¦‚æœæœ‰exit_timestampä¸”ä¸ä¸ºNoneå°±ç”¨å®ƒï¼Œå¦åˆ™ç”¨å½“å‰æ—¶é—´æˆ³
            signal_timestamp = signal.get('exit_timestamp') if signal.get('exit_timestamp') is not None else timestamp
            
            # ğŸ”§ å®‰å…¨è·å–ä»·æ ¼ï¼šä¼˜å…ˆä½¿ç”¨ priceï¼Œå…¶æ¬¡ä½¿ç”¨ new_stop_lossã€current_priceï¼Œæœ€åä½¿ç”¨å½“å‰æ”¶ç›˜ä»·
            signal_price = signal.get('price') or signal.get('new_stop_loss') or signal.get('current_price') or close_price
            
            trade_info = {
                'timestamp': signal_timestamp,
                'signal_type': signal['type'],
                'price': signal_price,
                'reason': signal.get('reason', '')
            }
            
            # æ·»åŠ æ­¢æŸæ­¢ç›ˆä¿¡æ¯ï¼ˆåªæœ‰å¼€ä»“ä¿¡å·æ‰æœ‰ï¼‰
            if 'stop_loss' in signal:
                trade_info['stop_loss'] = signal['stop_loss']
            if 'take_profit' in signal:
                trade_info['take_profit'] = signal['take_profit']
            
            # ğŸ”„ æ·»åŠ å¤æŠ•ç›¸å…³ä¿¡æ¯
            if 'invested_amount' in signal:
                trade_info['invested_amount'] = signal['invested_amount']
            if 'position_shares' in signal:
                trade_info['position_shares'] = signal['position_shares']
            if 'exit_amount' in signal:
                trade_info['exit_amount'] = signal['exit_amount']
            if 'cash_balance' in signal:
                trade_info['cash_balance'] = signal['cash_balance']
            if 'old_balance' in signal:
                trade_info['old_balance'] = signal['old_balance']
            if 'new_balance' in signal:
                trade_info['new_balance'] = signal['new_balance']
            
            if 'profit_loss' in signal:
                trade_info['profit_loss'] = signal['profit_loss']

            if 'transaction_fee' in signal:
                trade_info['transaction_fee'] = signal['transaction_fee']
            
            trades.append(trade_info)
            
            # ğŸ¨ æ”¶é›†äº¤æ˜“ä¿¡å·ï¼ˆç”¨äºå›¾è¡¨æ˜¾ç¤ºï¼‰
            chart_signal = {
                'time': signal_timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'type': signal['type'],
                'price': round(signal_price, 2),
                'reason': signal.get('reason', '')
            }
            chart_data['tradeSignals'].append(chart_signal)
            
            # ğŸ” è°ƒè¯•ä¿¡æ¯ï¼šéªŒè¯å¤æŠ•æ•°æ®ä¼ é€’
            if signal['type'] in ['OPEN_LONG', 'OPEN_SHORT']:
                invested = trade_info.get('invested_amount', 'N/A')
                balance = trade_info.get('cash_balance', 'N/A')
                print(f"    ğŸ”„ å¼€ä»“æ•°æ®éªŒè¯: æŠ•å…¥é‡‘é¢={invested} | ç°é‡‘ä½™é¢={balance}")
            elif signal['type'] in ['TAKE_PROFIT_LONG', 'TAKE_PROFIT_SHORT', 'STOP_LOSS_LONG', 'STOP_LOSS_SHORT']:
                old_balance = trade_info.get('old_balance', 'N/A')
                new_balance = trade_info.get('new_balance', 'N/A')
                print(f"    ğŸ”„ å¹³ä»“æ•°æ®éªŒè¯: æ—§ä½™é¢={old_balance} | æ–°ä½™é¢={new_balance}")
            
            # æ›´æ–°èµ„é‡‘çŠ¶æ€å’Œè¯¦ç»†ä¿¡æ¯
            if signal['type'] == 'OPEN_LONG':
                current_position = 'long'
                stop_loss = signal.get('stop_loss', 0)
                take_profit = signal.get('take_profit', None)
                risk = abs(signal_price - stop_loss) if stop_loss else 0
                
                if take_profit is not None:
                    reward = abs(take_profit - signal_price)
                    risk_reward_ratio = reward / risk if risk > 0 else 0
                    print(f"    ğŸ’° èµ„é‡‘: ${initial_capital:,.2f} | é£é™©: ${risk:.2f} | é¢„æœŸæ”¶ç›Š: ${reward:.2f} | é£é™©æ”¶ç›Šæ¯”: 1:{risk_reward_ratio:.2f}")
                else:
                    print(f"    ğŸ’° èµ„é‡‘: ${initial_capital:,.2f} | é£é™©: ${risk:.2f} | åŒå‘¨æœŸåŠ¨æ€æ­¢æŸç­–ç•¥ (æ— å›ºå®šæ­¢ç›ˆç›®æ ‡)")
                    
            elif signal['type'] == 'OPEN_SHORT':
                current_position = 'short'
                stop_loss = signal.get('stop_loss', 0)
                take_profit = signal.get('take_profit', None)
                risk = abs(stop_loss - signal_price) if stop_loss else 0
                
                if take_profit is not None:
                    reward = abs(signal_price - take_profit)
                    risk_reward_ratio = reward / risk if risk > 0 else 0
                    print(f"    ğŸ’° èµ„é‡‘: ${initial_capital:,.2f} | é£é™©: ${risk:.2f} | é¢„æœŸæ”¶ç›Š: ${reward:.2f} | é£é™©æ”¶ç›Šæ¯”: 1:{risk_reward_ratio:.2f}")
                else:
                    print(f"    ğŸ’° èµ„é‡‘: ${initial_capital:,.2f} | é£é™©: ${risk:.2f} | åŒå‘¨æœŸåŠ¨æ€æ­¢æŸç­–ç•¥ (æ— å›ºå®šæ­¢ç›ˆç›®æ ‡)")
            elif signal['type'] in ['STOP_LOSS_LONG', 'STOP_LOSS_SHORT', 'TAKE_PROFIT_LONG', 'TAKE_PROFIT_SHORT']:
                current_position = None

        # æ¯10000ä¸ªæ•°æ®ç‚¹ä¿å­˜ä¸€æ¬¡çŠ¶æ€
        if index % 10000 == 0:
            print(f"å·²å¤„ç† {index + 1}/{len(df)} ä¸ªæ•°æ®ç‚¹")

    # ğŸ¨ ä¿å­˜æœ€åä¸€ä¸ªKçº¿
    if current_kline is not None:
        chart_data['timeData'].append(current_kline_start.strftime('%Y-%m-%d %H:%M:%S'))
        chart_data['klineData'].append([
            round(current_kline['open'], 2),
            round(current_kline['close'], 2),
            round(current_kline['low'], 2),
            round(current_kline['high'], 2)
        ])
        
        # SARç›¸å…³æ•°æ®å·²åˆ é™¤ï¼Œä¸å†æ”¶é›†

    # === å›æµ‹ç»“æœç»Ÿè®¡ ===
    print(f"\n" + "=" * 60)
    print("ğŸ“Š çº¯VIDYAç­–ç•¥å›æµ‹ç»“æœ")
    print("=" * 60)
    
    total_trades = len(trades)
    if total_trades > 0:
        # ç»Ÿè®¡äº¤æ˜“ç±»å‹
        long_opens = len([t for t in trades if t['signal_type'] == 'OPEN_LONG'])
        short_opens = len([t for t in trades if t['signal_type'] == 'OPEN_SHORT'])
        long_stops = len([t for t in trades if t['signal_type'] == 'STOP_LOSS_LONG'])
        short_stops = len([t for t in trades if t['signal_type'] == 'STOP_LOSS_SHORT'])
        long_profits = len([t for t in trades if t['signal_type'] == 'TAKE_PROFIT_LONG'])
        short_profits = len([t for t in trades if t['signal_type'] == 'TAKE_PROFIT_SHORT'])
        
        # è®¡ç®—ç›ˆäº
        total_pnl = sum([t.get('profit_loss', 0) for t in trades if 'profit_loss' in t])
        
        print(f"ğŸ“ˆ æ€»äº¤æ˜“æ¬¡æ•°: {total_trades}")
        print(f"ğŸŸ¢ å¼€å¤šæ¬¡æ•°: {long_opens}")
        print(f"ğŸ”´ å¼€ç©ºæ¬¡æ•°: {short_opens}")
        print(f"âœ… å¤šå¤´æ­¢ç›ˆ: {long_profits}")
        print(f"âœ… ç©ºå¤´æ­¢ç›ˆ: {short_profits}")
        print(f"âŒ å¤šå¤´æ­¢æŸ: {long_stops}")
        print(f"âŒ ç©ºå¤´æ­¢æŸ: {short_stops}")
        print(f"ğŸ’° æ€»ç›ˆäº: ${total_pnl:,.2f}")
        
        # æ­¢ç›ˆæ­¢æŸæ¯”ä¾‹
        total_closes = long_stops + short_stops + long_profits + short_profits
        if total_closes > 0:
            profit_rate = (long_profits + short_profits) / total_closes * 100
            loss_rate = (long_stops + short_stops) / total_closes * 100
            print(f"ğŸ“Š æ­¢ç›ˆç‡: {profit_rate:.1f}% | æ­¢æŸç‡: {loss_rate:.1f}%")
        
        # æ˜¾ç¤ºæœ€è¿‘å‡ ç¬”äº¤æ˜“
        print(f"\nğŸ“‹ æœ€è¿‘5ç¬”äº¤æ˜“:")
        for trade in trades[-5:]:
            timestamp_str = trade['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if trade.get('timestamp') else 'N/A'
            pnl_str = f", ç›ˆäº: ${trade['profit_loss']:,.2f}" if 'profit_loss' in trade else ""
            print(f"  {timestamp_str} | {trade['signal_type']} | ä»·æ ¼: ${trade['price']:,.2f}{pnl_str}")
            print(f"    åŸå› : {trade['reason']}")
    else:
        print("âŒ æœªäº§ç”Ÿä»»ä½•äº¤æ˜“ä¿¡å·")
    
    # æ˜¾ç¤ºæœ€ç»ˆç­–ç•¥çŠ¶æ€
    final_status = strategy.get_current_status()
    print(f"\nğŸ¯ æœ€ç»ˆç­–ç•¥çŠ¶æ€:")
    print(f"  æŒä»“çŠ¶æ€: {final_status['position'] or 'ç©ºä»“'}")
    if final_status['entry_price']:
        print(f"  å…¥åœºä»·æ ¼: ${final_status['entry_price']:,.2f}")
    if final_status['stop_loss_level']:
        print(f"  æ­¢æŸä½: ${final_status['stop_loss_level']:,.2f}")
    if final_status['take_profit_level']:
        print(f"  æ­¢ç›ˆä½: ${final_status['take_profit_level']:,.2f}")
    # SARå€¼å·²åˆ é™¤
    
    print(f"\nğŸ’° èµ„é‡‘çŠ¶æ€:")
    print(f"  åˆå§‹èµ„é‡‘: ${initial_capital:,.2f} (å•è´¦æˆ·æ¨¡å¼)")
    print(f"  ç­–ç•¥ç°é‡‘ä½™é¢: ${strategy.cash_balance:,.2f}")
    
    # ğŸ” æ£€æŸ¥æŒä»“çŠ¶æ€
    if strategy.position is not None:
        print(f"  âš ï¸  è­¦å‘Šï¼šç­–ç•¥è¿˜æŒæœ‰ {strategy.position} ä»“ä½ï¼")
        print(f"      å…¥åœºä»·: ${strategy.entry_price:,.2f}")
        print(f"      æŒä»“ä»½é¢: {strategy.position_shares:.4f}")
        print(f"      æŠ•å…¥é‡‘é¢: ${strategy.current_invested_amount:,.2f}")
    else:
        print(f"  âœ… ç­–ç•¥å·²æ¸…ä»“")
    
    # === ç»©æ•ˆåˆ†æ ===
    print(f"\n" + "=" * 60)
    print("ğŸ“ˆ ç»©æ•ˆåˆ†æ")
    print("=" * 60)
    
    # åˆ›å»ºç»©æ•ˆåˆ†æå™¨
    analyzer = PerformanceAnalyzer(config)
    
    # å‡†å¤‡ä»·æ ¼æ•°æ® (timestamp, open, high, low, close)
    price_data = []
    for _, row in df.iterrows():
        price_data.append((row['timestamp'], row['open'], row['high'], row['low'], row['close']))
    
    # è®¡ç®—æ¯æ—¥å‡€å€¼
    nav_df = analyzer.calculate_daily_nav(trades, price_data)
    
    if not nav_df.empty:
        # è®¡ç®—å›æ’¤
        nav_df = analyzer.calculate_drawdown(nav_df)
        
        # è®¡ç®—å…³é”®ç»©æ•ˆæŒ‡æ ‡
        metrics = analyzer.calculate_performance_metrics(nav_df, trades)
        
        if metrics:
            # åˆ›å»ºç»Ÿä¸€çš„è¾“å‡ºç›®å½•
            output_dir = create_output_directory(config, metrics['annualized_return'])
            
            print(f"\nğŸ“Š å…³é”®ç»©æ•ˆæŒ‡æ ‡:")
            print(f"  ğŸ’° æ€»æ”¶ç›Šç‡: {metrics['total_return']:+.2f}%")
            print(f"  ğŸ“ˆ å¹´åŒ–æ”¶ç›Šç‡: {metrics['annualized_return']:+.2f}%")
            print(f"  ğŸ“‰ å¹´åŒ–æ³¢åŠ¨ç‡: {metrics['volatility']:.2f}%")
            print(f"  ğŸ¯ å¤æ™®æ¯”ç‡: {metrics['sharpe_ratio']:.2f}")
            print(f"  âš ï¸  æœ€å¤§å›æ’¤: {metrics['max_drawdown']:.2f}%")
            print(f"  ğŸ¯ èƒœç‡: {metrics['win_rate']:.1f}%")
            print(f"  ğŸ“Š ç›ˆäºæ¯”: {metrics['profit_loss_ratio']:.2f}")
            print(f"  ğŸ”„ æ€»äº¤æ˜“æ¬¡æ•°: {metrics['total_trades']}")
            print(f"  ğŸ“… äº¤æ˜“å¤©æ•°: {metrics['trading_days']}")
            print(f"  ğŸ’ æœ€ç»ˆå‡€å€¼: ${metrics['final_nav']:,.2f}")
            
            # === ç”Ÿæˆæ‰€æœ‰ç»“æœæ–‡ä»¶åˆ°ç»Ÿä¸€ç›®å½• ===
            print(f"\n" + "=" * 60)
            print("ğŸ“ ç”Ÿæˆç»“æœæ–‡ä»¶")
            print("=" * 60)
            
            # 1. å¯¼å‡ºäº¤æ˜“è®°å½•åˆ°Excel
            try:
                excel_file = export_trades_to_excel(trades, config, output_dir)
                if excel_file:
                    print(f"ğŸ“‹ äº¤æ˜“è®°å½•: {os.path.abspath(excel_file)}")
            except Exception as e:
                print(f"âš ï¸  äº¤æ˜“è®°å½•å¯¼å‡ºå¤±è´¥: {e}")
            
            # 2. ç”Ÿæˆç»©æ•ˆå›¾è¡¨
            try:
                chart_files = analyzer.generate_performance_charts(nav_df, output_dir)
                print(f"ğŸ“ˆ ç»©æ•ˆå›¾è¡¨å·²ç”Ÿæˆ")
            except Exception as e:
                print(f"âš ï¸  ç»©æ•ˆå›¾è¡¨ç”Ÿæˆå¤±è´¥: {e}")
            
            # 3. ç”Ÿæˆç»©æ•ˆExcelæŠ¥å‘Š
            try:
                perf_excel = analyzer.generate_performance_excel(nav_df, output_dir)
                if perf_excel:
                    print(f"ğŸ“‹ ç»©æ•ˆExcelæŠ¥å‘Š: {os.path.abspath(perf_excel)}")
            except Exception as e:
                print(f"âš ï¸  ç»©æ•ˆExcelæŠ¥å‘Šç”Ÿæˆå¤±è´¥: {e}")
            
            # 4. ç”ŸæˆHTMLæŠ¥å‘Š
            try:
                html_report = analyzer.generate_html_report(nav_df, config['long_coin'], trades, config, output_dir)
                if html_report:
                    print(f"ğŸŒ HTMLç»©æ•ˆæŠ¥å‘Š: {os.path.abspath(html_report)}")
            except Exception as e:
                print(f"âš ï¸  HTMLæŠ¥å‘Šç”Ÿæˆå¤±è´¥: {e}")
            
            # 5. ç”Ÿæˆäº¤äº’å¼å›¾è¡¨
            try:
                chart_html = create_chart_html(chart_data, output_dir)
                if chart_html:
                    print(f"ğŸ¨ äº¤äº’å¼å›¾è¡¨: {os.path.abspath(chart_html)}")
            except Exception as e:
                print(f"âš ï¸  äº¤äº’å¼å›¾è¡¨ç”Ÿæˆå¤±è´¥: {e}")
            
            print(f"\nâœ… æ‰€æœ‰ç»“æœæ–‡ä»¶å·²ä¿å­˜åˆ°: {os.path.abspath(output_dir)}")
    else:
        print("âš ï¸  æ— æ³•è®¡ç®—å‡€å€¼ï¼Œè·³è¿‡ç»©æ•ˆåˆ†æ")
    
    # å…³é—­æ•°æ®åº“è¿æ¥
    db_service.disconnect()
    
    print(f"\nğŸ‰ å›æµ‹å®Œæˆï¼æ‰€æœ‰ç»“æœæ–‡ä»¶å·²æŒ‰å¸ç§å’Œå¹´åŒ–æ”¶ç›Šç‡åˆ†ç±»ä¿å­˜")

if __name__ == "__main__":
    main() 